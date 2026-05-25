#!/usr/bin/env python3
"""
Phase 19.3 — Re-theme the two sellable Excel products to match the
Direction E brand palette (indigo / apricot / cream / paper).

What it changes
  - Solid fills: old palette hex → Direction E hex (cell by cell)
  - Font colors: same swap
  - Border colors: same swap
  - Sheet tab colors: apricot for input-heavy sheets ("Setup",
    "Trip Log", "Income Forecast", "Tax Calculator", "Quarterly
    Payments"), indigo for all reference/output sheets.

What it does NOT change
  - Formulas, validations, dropdowns — preserved exactly
  - Fonts (Calibri is the existing template convention — per xlsx
    skill guidance, existing conventions override; user only asked
    for "color/theme")
  - Row/column structure, merged ranges, sheet order

Idempotent: re-running maps remap targets to themselves.

Run from repo root:  python3 scripts/rebrand-xlsx.py
"""
from pathlib import Path
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side

ROOT = Path(__file__).resolve().parent.parent
PRODUCTS = ROOT / "products"

# Direction E brand tokens (ARGB hex — openpyxl prefers full alpha 'FF')
INDIGO  = "FF2D3068"
APRICOT = "FFE89464"
CREAM   = "FFF0ECE1"
PAPER   = "FFFBF8EE"
INDIGO_55 = "FF5F6385"   # muted body / secondary text
WARN      = "FFC98A3A"   # Direction E --warn (amber-gold accent)
WARN_BG   = "FFF5EBD8"   # --warn-bg flattened over cream FBF8EE @ 12% alpha

# Old palette → new palette.
# Both 6-char and 8-char ARGB variants because openpyxl returns either
# depending on whether the style was set with or without alpha.
COLOR_MAP = {
    "FF1C2B4A": INDIGO,   "1C2B4A": INDIGO[2:],
    "FFFDF5E8": PAPER,    "FDF5E8": PAPER[2:],
    "FFF0ECE3": CREAM,    "F0ECE3": CREAM[2:],
    "FFFAF8F4": PAPER,    "FAF8F4": PAPER[2:],
    "FF6B7A96": INDIGO_55,"6B7A96": INDIGO_55[2:],
    # Off-brand accents found mid-rebrand: amber-gold header + pink
    # background tint. Both used for "attention / amount owed" cells.
    # Map to Direction E --warn family for the same semantic meaning.
    "FFC9973A": WARN,     "C9973A": WARN[2:],
    "FFFBE9E9": WARN_BG,  "FBE9E9": WARN_BG[2:],
}

INPUT_SHEETS = {
    "Setup", "Trip Log", "Income Forecast",
    "Tax Calculator", "Quarterly Payments",
}


def remap(rgb):
    """Return the new hex if rgb is in COLOR_MAP, else rgb unchanged."""
    if not isinstance(rgb, str):
        return rgb  # could be a theme index, RGB tuple, or None
    return COLOR_MAP.get(rgb.upper(), rgb)


def restyle_fill(cell):
    if cell.fill.patternType != "solid":
        return False
    old = cell.fill.start_color.rgb
    new = remap(old)
    if new == old:
        return False
    cell.fill = PatternFill("solid", start_color=new, end_color=new)
    return True


def restyle_font(cell):
    if cell.font.color is None or cell.font.color.rgb is None:
        return False
    old = cell.font.color.rgb
    new = remap(old)
    if new == old:
        return False
    # Copy() preserves name/size/bold/italic/underline etc.
    new_font = copy(cell.font)
    new_font.color = new
    cell.font = new_font
    return True


def restyle_border(cell):
    """Walk all 4 sides + diagonal; remap each side's color."""
    b = cell.border
    if not b:
        return False
    changed = False
    sides = {}
    for name in ("left", "right", "top", "bottom", "diagonal"):
        side = getattr(b, name, None)
        if side and side.color and isinstance(side.color.rgb, str):
            new = remap(side.color.rgb)
            if new != side.color.rgb:
                sides[name] = Side(border_style=side.style, color=new)
                changed = True
    if not changed:
        return False
    # Rebuild border preserving unchanged sides
    cell.border = Border(
        left=sides.get("left", b.left),
        right=sides.get("right", b.right),
        top=sides.get("top", b.top),
        bottom=sides.get("bottom", b.bottom),
        diagonal=sides.get("diagonal", b.diagonal),
        diagonal_direction=b.diagonal_direction,
        outline=b.outline,
        vertical=b.vertical,
        horizontal=b.horizontal,
    )
    return True


def restyle(path: Path) -> dict:
    wb = load_workbook(path)
    counters = {"fills": 0, "fonts": 0, "borders": 0, "tabs": 0}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Tab color: apricot for input sheets, indigo otherwise
        tab = APRICOT if sheet_name in INPUT_SHEETS else INDIGO
        ws.sheet_properties.tabColor = tab[2:]  # tabColor uses 6-char hex
        counters["tabs"] += 1

        for row in ws.iter_rows():
            for cell in row:
                if restyle_fill(cell):   counters["fills"]  += 1
                if restyle_font(cell):   counters["fonts"]  += 1
                if restyle_border(cell): counters["borders"] += 1

    wb.save(path)
    return counters


def main() -> int:
    files = sorted(PRODUCTS.glob("*.xlsx"))
    if not files:
        print(f"No .xlsx files found in {PRODUCTS}")
        return 1
    for f in files:
        c = restyle(f)
        print(f"  [{f.name}]  fills={c['fills']:>3}  fonts={c['fonts']:>3}  "
              f"borders={c['borders']:>3}  tabs={c['tabs']:>2}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
