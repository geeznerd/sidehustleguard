#!/usr/bin/env python3
"""
Polish pass on the STR Host Tax Tracker. Earlier rebrand_xlsx.py changed
colors + fonts but the user said the result still 'looks sloppy'.

Real polish moves (typography + color alone don't carry it):

  1. Hide default gridlines on every sheet — the single biggest 'this is
     a template, not a spreadsheet' signal Excel can give.

  2. Alignment overhaul:
       - Text columns left-aligned (Property / Notes / Vendor / Category)
       - Numeric columns right-aligned (currency / % / dates)
       - Column headers stay center-aligned
     Center-aligning everything was the default and reads as 'default'.

  3. Row height bumps for breathing room:
       - Header bands: +20% (32 → 38, 26 → 32)
       - Data rows: +30% (20 → 26)
     Inter + Fraunces both have taller x-heights than Calibri at the same
     point size; the originals were sized for Calibri.

  4. Subtle horizontal rhythm — replace heavy borders with a single
     1px bottom hairline (indigo-08) on data rows. Most data tables had
     ALL FOUR sides bordered, which reads heavy.

  5. Freeze panes on the data sheets so column headers stay visible
     while scrolling.

  6. Remove stray fills from empty cells that aren't part of a deliberate
     header band — these accumulated from the original template and clutter
     the visual rhythm.

  7. Reduce the apricot-pale 'example row' fill saturation so the example
     row in the Booking Log + Deductions doesn't compete with real data.
"""
from __future__ import annotations
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Color, Font
from openpyxl.utils import get_column_letter

SRC = Path('/Users/dork/Downloads/STR-Host-Tax-Tracker-2026-DirectionE-final.xlsx')
DST = Path('/Users/dork/Downloads/STR-Host-Tax-Tracker-2026-DirectionE-polished.xlsx')

INDIGO_08 = '00E6E2D6'      # warm hairline color (cream + ~8% indigo)
APRICOT_PALE_SOFT = '00FDF1E5'  # softer than #FDE8D5 — barely-there tint


def is_numeric_format(fmt: str) -> bool:
    """Return True if a number_format string represents a numeric value."""
    if not fmt or fmt == 'General':
        return False
    return any(t in fmt for t in ['$', '#', '0', '%', 'YYYY', 'MM', 'DD'])


def polish_sheet(ws):
    # 1. Hide gridlines
    ws.sheet_view.showGridLines = False

    # Identify the "header row" of each data table — usually row 4 (after the
    # title band at row 1, KPI band at row 2, gap at row 3). Sheets like
    # Quarterly Est. and Instructions have different layouts; we'll only
    # apply table polish if there's a clear header band at row 4.
    has_table_header = False
    if ws.max_row >= 4:
        row4 = list(ws[4])
        # If most cells in row 4 have indigo fill, treat it as a header
        indigo_count = sum(1 for c in row4
                            if c.fill.patternType and
                            getattr(c.fill.fgColor, 'rgb', None) == '002D3068')
        has_table_header = indigo_count >= 5

    hairline = Side(style='thin', color=INDIGO_08)
    no_side = Side(style=None)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            try:
                # 2. Alignment overhaul — format takes priority over Python type
                # (dates can be stored as strings like '2026-01-15' but should
                # still right-align like numbers).
                fmt = cell.number_format or 'General'
                cur = cell.alignment
                if cell.row >= 5 and cell.value is not None:
                    # Decide horizontal alignment
                    if is_numeric_format(fmt) or isinstance(cell.value, (int, float)):
                        # number / date / % cell — right
                        horiz = 'right'
                    elif isinstance(cell.value, str) and cell.value.startswith('='):
                        # formula — fall back to format hint, default right
                        horiz = 'right' if is_numeric_format(fmt) else 'left'
                    else:
                        # text cell — left
                        horiz = 'left'
                    cell.alignment = Alignment(
                        horizontal=horiz,
                        vertical='center',
                        wrap_text=cur.wrap_text,
                        indent=1 if cur.indent == 0 else cur.indent,
                    )
                # Column headers (row 4 with indigo fill): center, vertical center
                elif cell.row == 4 and has_table_header:
                    cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center',
                        wrap_text=True,
                    )

                # 4. Subtle hairline borders — replace heavy borders on data rows
                if cell.row >= 5 and has_table_header and cell.value is not None:
                    cell.border = Border(top=no_side, bottom=hairline,
                                         left=no_side, right=no_side)
                elif cell.row == 4 and has_table_header:
                    # Header row: top + bottom indigo hairlines, no sides
                    cell.border = Border(top=no_side, bottom=hairline,
                                         left=no_side, right=no_side)

                # 6. Remove stray fills from empty cells that aren't part of
                # the deliberate indigo header band (row 1 + row 4)
                if cell.value is None and cell.row not in (1, 2, 4):
                    try:
                        rgb = cell.fill.fgColor.rgb
                        # Strip apricot-pale fills from empty cells (leftover
                        # from example-row continuation that doesn't carry data)
                        if rgb in ('00FDE8D5', '00FDF1E5') and cell.row > 5:
                            cell.fill = PatternFill(fill_type=None)
                    except Exception:
                        pass

                # 7. Soften the example-row apricot fill if it's the only
                # apricot-pale fill (i.e. the example row, row 5)
                if cell.row == 5 and cell.value is not None:
                    try:
                        rgb = cell.fill.fgColor.rgb
                        if rgb == '00FDE8D5':
                            cell.fill = PatternFill(
                                patternType='solid',
                                fgColor=Color(rgb=APRICOT_PALE_SOFT),
                            )
                    except Exception:
                        pass

            except Exception:
                continue

    # 3. Row height bumps
    # Heuristic: tables have row 4 as header → bump rows 4+ to 26pt minimum
    if has_table_header:
        if ws.row_dimensions[4].height and ws.row_dimensions[4].height < 36:
            ws.row_dimensions[4].height = 36
        for r in range(5, ws.max_row + 1):
            h = ws.row_dimensions[r].height
            # Skip explicit small/spacer rows
            if h is not None and h < 14:
                continue
            ws.row_dimensions[r].height = max(h or 20, 26)
    # Title row + KPI band → bump
    if ws.row_dimensions[1].height and ws.row_dimensions[1].height < 56:
        # Only bump if row 1 has the brand indigo band
        cell = ws.cell(row=1, column=2)
        try:
            if cell.fill.patternType and cell.fill.fgColor.rgb == '002D3068':
                ws.row_dimensions[1].height = max(ws.row_dimensions[1].height or 0, 56)
        except Exception:
            pass

    # 5. Freeze pane below column headers (row 4 is data-table header on most sheets)
    if has_table_header:
        ws.freeze_panes = 'A5'


def polish(src: Path, dst: Path):
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst)
    for sname in wb.sheetnames:
        ws = wb[sname]
        polish_sheet(ws)
    wb.save(dst)


if __name__ == '__main__':
    polish(SRC, DST)
    print(f'Wrote: {DST}')
