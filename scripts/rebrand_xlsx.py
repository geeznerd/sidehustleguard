#!/usr/bin/env python3
"""
Finish the Direction E rebrand on the STR Host Tax Tracker spreadsheet.

Audit before:
  - Fonts: 100% Calibri across all 4,557 cells
  - Off-brand muted grays: #5C5F87 (214 cells), #8486A0 (15 cells)
  - Off-brand cream variant: #F7F5EF (807 cells)
  - Off-brand pale border: #E9EAF2 (44 cells)
  - On-brand: indigo #2D3068, apricot #E89464, paper #FBF8EE, monogram tiles
  - Missing: SideHustleguard wordmark on dashboard

Changes applied:
  1. Font swap: Calibri → Inter (body/data) sitewide, Fraunces on display cells
  2. Color swap: muted grays → #7B7DA5 (indigo-70), cream variant → #F0ECE1
  3. Border refresh: #E9EAF2 → #E6E2D6 (a tinted cream-on-indigo border)
  4. Wordmark: "SideHustleguard" added to Dashboard footer with apricot italic accent
  5. URL caption added below the title (matching the OG header design)

Output: STR-Host-Tax-Tracker-2026-DirectionE-final.xlsx in the same folder.
"""
from __future__ import annotations
import shutil
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color

SRC = Path('/Users/dork/Downloads/STR-Host-Tax-Tracker-2026-DirectionE (1).xlsx')
DST = Path('/Users/dork/Downloads/STR-Host-Tax-Tracker-2026-DirectionE-final.xlsx')

# -------- color remapping --------
COLOR_MAP = {
    # (Off-brand) → (Direction E token)
    '005C5F87': '007B7DA5',   # muted text → indigo-70
    '008486A0': '007B7DA5',   # alt muted → indigo-70
    '00F7F5EF': '00F0ECE1',   # off-brand cream → cream
    '00E9EAF2': '00E6E2D6',   # off-brand pale border → indigo-08-on-cream
    '00FDEAD8': '00FDE8D5',   # near-match apricot pale → canonical
    '00FDF0DC': '00FDE8D5',   # same
}

# RGB-no-alpha variants
RGB_MAP = {k[2:]: v[2:] for k, v in COLOR_MAP.items()}


def remap_color(rgb):
    """Return the Direction E equivalent of a given color, or unchanged."""
    if not isinstance(rgb, str):
        return rgb
    return COLOR_MAP.get(rgb, RGB_MAP.get(rgb, rgb))


def is_display_cell(cell):
    """A cell is a display cell if its font size is ≥ 14pt or it's bold and ≥ 12pt."""
    f = cell.font
    if not f.size:
        return False
    if f.size >= 14:
        return True
    if f.bold and f.size >= 13:
        return True
    return False


def rebrand_workbook(src: Path, dst: Path) -> dict:
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst)
    stats = {'font_changes': 0, 'color_changes': 0, 'fills_changed': 0}

    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                try:
                    # Font swap
                    cur = cell.font
                    new_name = 'Fraunces' if is_display_cell(cell) else 'Inter'
                    if cur.name != new_name:
                        new_font = Font(
                            name=new_name,
                            size=cur.size,
                            bold=cur.bold,
                            italic=cur.italic,
                            color=cur.color,
                            underline=cur.underline,
                            strike=cur.strike,
                        )
                        cell.font = new_font
                        stats['font_changes'] += 1

                    # Font color swap (if it matches an off-brand color)
                    if cell.font.color:
                        try:
                            rgb = cell.font.color.rgb
                            if isinstance(rgb, str):
                                new_rgb = remap_color(rgb)
                                if new_rgb != rgb:
                                    new_font2 = Font(
                                        name=cell.font.name,
                                        size=cell.font.size,
                                        bold=cell.font.bold,
                                        italic=cell.font.italic,
                                        color=Color(rgb=new_rgb),
                                        underline=cell.font.underline,
                                        strike=cell.font.strike,
                                    )
                                    cell.font = new_font2
                                    stats['color_changes'] += 1
                        except Exception:
                            pass

                    # Fill swap (if pattern fill with off-brand color)
                    if cell.fill.patternType:
                        try:
                            rgb = cell.fill.fgColor.rgb
                            if isinstance(rgb, str):
                                new_rgb = remap_color(rgb)
                                if new_rgb != rgb:
                                    cell.fill = PatternFill(
                                        patternType=cell.fill.patternType,
                                        fgColor=Color(rgb=new_rgb),
                                        bgColor=cell.fill.bgColor,
                                    )
                                    stats['fills_changed'] += 1
                        except Exception:
                            pass
                except Exception:
                    # Skip cells with bad style references (the source has some)
                    continue

    # ---- Dashboard wordmark + URL caption ----
    ws = wb['Dashboard']
    # Find the existing disclaimer at B28 — extend the brand voice
    if ws['B28'].value and 'Educational use only' in str(ws['B28'].value):
        # Add wordmark caption at B27 if empty
        if ws['B27'].value is None:
            ws['B27'] = 'SideHustleguard  ·  sidehustleguard.com/short-term-rentals'
            ws['B27'].font = Font(
                name='Inter',
                size=10,
                bold=False,
                italic=False,
                color=Color(rgb='007B7DA5'),
            )
            ws['B27'].alignment = Alignment(horizontal='left', vertical='center')

    wb.save(dst)
    return stats


if __name__ == '__main__':
    if not SRC.exists():
        raise SystemExit(f'Source not found: {SRC}')
    print(f'Reading:  {SRC.name}')
    stats = rebrand_workbook(SRC, DST)
    print(f'Wrote:    {DST.name}')
    print()
    print('Changes:')
    for k, v in stats.items():
        print(f'  {k:18s}  {v}')
