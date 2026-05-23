#!/usr/bin/env python3
"""
Fix the three concrete bugs visible in the user's dashboard screenshot:

  1. HTML entity literals — '&amp;' showing as text instead of '&':
       - B3 (subtitle):  '... deductions &amp; quarterly estimates ...'
       - B12 (caption):  'Before fees &amp; TOT'

  2. "Open →" overflow on JUMP TO cards. The title cells are merged
     D20:H20 / D22:H22 / L20:P20 / L22:P22 / D25:P25 — leaving the
     "Open →" cells at single-column width (13pt). The arrow visibly
     bleeds past the cell edge.
     Fix: shorten title merges by one column and extend the "Open" cells
     by one column so they're 2 cells wide (26pt).

  3. "Note: STR (Sch E) = $0" overflow in the Est. Tax (Federal) metric.
     Cell J11 is merged J11:K11 (26pt wide) and the note string is too long.
     Fix: shorten the source string in Quarterly Est. C21 to fit, AND enable
     wrap_text so it can break across two lines if still tight.
"""
from __future__ import annotations
import shutil
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment

SRC = Path('/Users/dork/Downloads/STR-Host-Tax-Tracker-2026-DirectionE-polished.xlsx')
DST = Path('/Users/dork/Downloads/STR-Host-Tax-Tracker-2026-DirectionE-final-v2.xlsx')

shutil.copy(SRC, DST)
wb = openpyxl.load_workbook(DST)

# ---- 1. HTML entity literals ----
ws = wb['Dashboard']
ws['B3'] = 'Short-Term Rental income, TOT, deductions & quarterly estimates — all in one place'
ws['B12'] = 'Before fees & TOT'

# Also sweep all sheets for any remaining '&amp;'
fixed = 0
for sname in wb.sheetnames:
    s = wb[sname]
    for row in s.iter_rows():
        for c in row:
            if isinstance(c.value, str) and '&amp;' in c.value:
                c.value = c.value.replace('&amp;', '&')
                fixed += 1
print(f'Entity fixes total: {fixed}')

# ---- 2. JUMP TO "Open →" overflow ----
ws = wb['Dashboard']
# Unmerge the existing title ranges, then re-merge with one fewer col,
# then unmerge the single-cell "Open" range and re-merge spanning two cols.
def safe_unmerge(ws, range_str):
    """Unmerge a range if it's currently merged."""
    targets = [m for m in list(ws.merged_cells.ranges) if str(m) == range_str]
    for m in targets:
        ws.unmerge_cells(str(m))

# Save the "Open →" anchor cell values BEFORE any merges (merges clear non-anchor cells)
def open_text_at(coord, default='Open →'):
    """Capture the current 'Open →' value before we destroy it with merges."""
    v = ws[coord].value
    return v if v else default

i20 = open_text_at('I20')
i22 = open_text_at('I22')
q20 = open_text_at('Q20')
q22 = open_text_at('Q22')
q25 = open_text_at('Q25')

# Booking Log row 20: title D20:H20 → D20:G20 ; Open at H20:I20 (anchor H20)
safe_unmerge(ws, 'D20:H20')
ws['I20'] = None  # explicit clear to avoid old value bleeding into new merge
ws.merge_cells('D20:G20')
ws.merge_cells('H20:I20')
ws['H20'] = i20

# Deductions row 22
safe_unmerge(ws, 'D22:H22')
ws['I22'] = None
ws.merge_cells('D22:G22')
ws.merge_cells('H22:I22')
ws['H22'] = i22

# TOT Tracker row 20: title L20:P20 → L20:O20 ; Open at P20:Q20 (anchor P20)
safe_unmerge(ws, 'L20:P20')
ws['Q20'] = None
ws.merge_cells('L20:O20')
ws.merge_cells('P20:Q20')
ws['P20'] = q20

# Quarterly Est. row 22
safe_unmerge(ws, 'L22:P22')
ws['Q22'] = None
ws.merge_cells('L22:O22')
ws.merge_cells('P22:Q22')
ws['P22'] = q22

# Instructions row 25: title D25:P25 → D25:O25 ; Open at P25:Q25 (anchor P25)
safe_unmerge(ws, 'D25:P25')
ws['Q25'] = None
ws.merge_cells('D25:O25')
ws.merge_cells('P25:Q25')
ws['P25'] = q25

# Right-align the Open → cells so the arrow points outward cleanly
for coord in ['H20', 'H22', 'P20', 'P22', 'P25']:
    c = ws[coord]
    cur = c.alignment
    c.alignment = Alignment(
        horizontal='right',
        vertical='center',
        wrap_text=False,
        indent=1,
    )

# ---- 3. "Note: STR (Sch E) = $0" overflow ----
# Shorten the source text and enable wrap on the metric cell + caption
ws_q = wb['Quarterly Est.']
if ws_q['C21'].value == 'Note: STR (Sch E) = $0':
    ws_q['C21'] = 'STR Sch E · $0'

# Enable wrap on the Dashboard metric cell J11:K11 so if any future text grows,
# it wraps cleanly instead of overflowing
ws = wb['Dashboard']
for coord in ['J11', 'B11', 'D11', 'F11', 'H11', 'L11']:
    c = ws[coord]
    c.alignment = Alignment(
        horizontal='right',
        vertical='center',
        wrap_text=True,
    )

wb.save(DST)
print(f'Wrote: {DST}')
