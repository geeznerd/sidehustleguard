"""
Build the Gig Driver Mileage Tracker 2026 — Excel/Google-Sheets-compatible product.
Outputs: gig-driver-mileage-tracker-2026.xlsx in the same directory.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# ── BRAND COLORS ──
NAVY      = "FF1C2B4A"
GOLD      = "FFC9973A"
GOLD_LT   = "FFDDB06A"
GOLD_PALE = "FFFDF5E8"
CREAM     = "FFFAF8F4"
CREAM_D   = "FFF0ECE3"
GREEN     = "FF276944"
GREEN_BG  = "FFEAF5EF"
AMBER     = "FFA86C0A"
AMBER_BG  = "FFFEF6E8"
WHITE     = "FFFFFFFF"
GRAY      = "FF6B7A96"
NAVY_DK   = "FF14213D"
BORDER    = "FFE5E7ED"

# ── HELPERS ──
def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def thin_box(color=BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_bottom(color=NAVY):
    return Border(bottom=Side(style="medium", color=color))

def cell(ws, coord, value, *, font=None, fillc=None, align=None, border=None, fmt=None):
    c = ws[coord]
    c.value = value
    if font: c.font = font
    if fillc: c.fill = fillc
    if align: c.alignment = align
    if border: c.border = border
    if fmt: c.number_format = fmt
    return c

# Reusable fonts
F_TITLE    = Font(name="Calibri", size=26, bold=True, color=NAVY)
F_SUBTITLE = Font(name="Calibri", size=11, italic=True, color=GOLD)
F_H1       = Font(name="Calibri", size=15, bold=True, color=NAVY)
F_H2       = Font(name="Calibri", size=12, bold=True, color=NAVY)
F_BODY     = Font(name="Calibri", size=11, color="FF333333")
F_MUTED    = Font(name="Calibri", size=10, color=GRAY)
F_TH       = Font(name="Calibri", size=11, bold=True, color=WHITE)
F_TOTAL    = Font(name="Calibri", size=12, bold=True, color=NAVY)
F_KPI_LBL  = Font(name="Calibri", size=10, bold=True, color=GRAY)
F_KPI_VAL  = Font(name="Calibri", size=22, bold=True, color=NAVY)
F_BIG_GOLD = Font(name="Calibri", size=22, bold=True, color=GOLD.replace("FF","FF"))

A_LEFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
A_LEFT_T= Alignment(horizontal="left",   vertical="top",    wrap_text=True)
A_CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_RIGHT = Alignment(horizontal="right",  vertical="center")

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════
# SHEET 1 — START HERE
# ════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Start Here"
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110

# Column widths
ws.column_dimensions["A"].width = 2
for col, w in zip("BCDEFGH", [22, 18, 18, 18, 18, 18, 18]):
    ws.column_dimensions[col].width = w

# Title block
ws.row_dimensions[2].height = 38
cell(ws, "B2", "Gig Driver Mileage Tracker 2026", font=F_TITLE, align=A_LEFT)
ws.merge_cells("B2:H2")

cell(ws, "B3", "SideHustleGuard · sidehustleguard.com", font=F_SUBTITLE, align=A_LEFT)
ws.merge_cells("B3:H3")

# Hero callout
ws.row_dimensions[5].height = 6
ws.row_dimensions[6].height = 22
cell(ws, "B6", "What this tracker does", font=F_H1, align=A_LEFT)
ws.merge_cells("B6:H6")

intro = ("Logs every business trip you take for DoorDash, Uber, Lyft, Grubhub, Instacart, "
         "Amazon Flex, TaskRabbit, Rover, or any other gig platform. Auto-calculates your "
         "IRS standard mileage deduction — the single biggest tax break gig workers get, "
         "typically worth $5,000–$15,000+ off your taxable income each year.")
ws.row_dimensions[7].height = 70
cell(ws, "B7", intro, font=F_BODY, align=A_LEFT_T)
ws.merge_cells("B7:H7")

# 3-step block
ws.row_dimensions[9].height = 22
cell(ws, "B9", "How to use it (3 steps)", font=F_H1, align=A_LEFT)
ws.merge_cells("B9:H9")

steps = [
    ("1.", "Open the Setup tab.", "Set the tax year, confirm the IRS mileage rate (we pre-fill it; verify at IRS.gov each January), and list your vehicles + the platforms you drive for."),
    ("2.", "Log trips on the Trip Log tab.", "One row per trip. Pick the date, vehicle, and platform from dropdowns. Enter miles directly OR enter your start and end odometer — whichever is faster. Default business % is 100 because most gig trips are 100% business."),
    ("3.", "Check the Monthly & Year-End Summary tabs.", "Everything totals automatically. The Year-End Summary tells you exactly what to write on Schedule C line 9 (Car & Truck Expenses)."),
]
row = 10
for num, head, body in steps:
    ws.row_dimensions[row].height = 56
    cell(ws, "B"+str(row), num, font=Font(name="Calibri", size=22, bold=True, color=GOLD), align=Alignment(horizontal="center", vertical="top"))
    cell(ws, "C"+str(row), head, font=F_H2, align=A_LEFT_T)
    cell(ws, "D"+str(row), body, font=F_BODY, align=A_LEFT_T)
    ws.merge_cells("D{0}:H{0}".format(row))
    row += 1

# IRS-compliance callout
row += 1
ws.row_dimensions[row].height = 22
cell(ws, "B"+str(row), "What the IRS actually requires", font=F_H1, align=A_LEFT)
ws.merge_cells("B"+str(row)+":H"+str(row))
row += 1

requirements = [
    "A contemporaneous log — entries made at or near the time of the trip, not reconstructed at year-end",
    "Date of each business trip",
    "Business purpose (e.g., \"DoorDash delivery,\" \"Uber pickup → drop-off\")",
    "Total miles driven for that trip (or start + end odometer)",
    "Total miles driven for the year, including personal — this is the denominator that proves your business %",
]
for item in requirements:
    ws.row_dimensions[row].height = 22
    cell(ws, "B"+str(row), "✓", font=Font(name="Calibri", size=12, bold=True, color=GREEN), align=A_CTR)
    cell(ws, "C"+str(row), item, font=F_BODY, align=A_LEFT)
    ws.merge_cells("C{0}:H{0}".format(row))
    row += 1

row += 1
ws.row_dimensions[row].height = 50
note = ("This tracker captures all of the above. Save the file (or a print/PDF export) with your tax records for at least 3 years after filing — that's the standard IRS audit lookback window.")
cell(ws, "B"+str(row), note, font=F_MUTED, align=A_LEFT_T)
ws.merge_cells("B{0}:H{0}".format(row))

# Disclaimer
row += 2
ws.row_dimensions[row].height = 50
disc = ("Educational tool only — not tax or legal advice. The IRS standard mileage rate is set annually; verify the current year's rate at irs.gov/standardmileagerates before filing. For complex situations (employer reimbursements, multiple businesses, leased vehicles), consult a CPA.")
cell(ws, "B"+str(row), disc, font=Font(name="Calibri", size=9, italic=True, color=GRAY), align=A_LEFT_T)
ws.merge_cells("B{0}:H{0}".format(row))

# ════════════════════════════════════════════════════════════════════
# SHEET 2 — SETUP
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Setup")
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 22

ws.row_dimensions[2].height = 32
cell(ws, "B2", "Setup", font=Font(name="Calibri", size=22, bold=True, color=NAVY), align=A_LEFT)
ws.merge_cells("B2:F2")

cell(ws, "B3", "Edit the gold cells below. The rest of the tracker reads from here.", font=F_MUTED, align=A_LEFT)
ws.merge_cells("B3:F3")

# Year + rate block
ws.row_dimensions[5].height = 24
cell(ws, "B5", "Tax year & rate", font=F_H1, align=A_LEFT)
ws.merge_cells("B5:F5")

# Field rows
def field(row_idx, label, value, fmt=None, is_input=True):
    cell(ws, f"B{row_idx}", label, font=F_BODY, align=A_LEFT)
    c = ws[f"C{row_idx}"]
    c.value = value
    c.font = Font(name="Calibri", size=12, bold=True, color=NAVY)
    c.alignment = A_LEFT
    c.fill = fill(GOLD_PALE) if is_input else fill(CREAM_D)
    c.border = thin_box(GOLD if is_input else BORDER)
    if fmt: c.number_format = fmt
    ws.row_dimensions[row_idx].height = 24

field(6, "Tax year", 2026)
field(7, "IRS standard mileage rate ($/mile)", 0.70, fmt='"$"#,##0.000')

cell(ws, "B8", "← Verify at irs.gov/standardmileagerates each January. The IRS publishes the rate for the calendar year in mid-December.", font=F_MUTED, align=A_LEFT)
ws.merge_cells("B8:F8")

# Named ranges so formulas elsewhere can reference these
wb.defined_names["TaxYear"]  = DefinedName("TaxYear",  attr_text="Setup!$C$6")
wb.defined_names["MileRate"] = DefinedName("MileRate", attr_text="Setup!$C$7")

# Your vehicles
ws.row_dimensions[10].height = 24
cell(ws, "B10", "Your vehicles", font=F_H1, align=A_LEFT)
ws.merge_cells("B10:F10")

ws.row_dimensions[11].height = 22
cell(ws, "B11", "Vehicle name", font=F_TH, fillc=fill(NAVY), align=A_LEFT)
cell(ws, "C11", "Year / Make / Model", font=F_TH, fillc=fill(NAVY), align=A_LEFT)
cell(ws, "D11", "Starting odometer (Jan 1)", font=F_TH, fillc=fill(NAVY), align=A_LEFT)
cell(ws, "E11", "Ending odometer (Dec 31)", font=F_TH, fillc=fill(NAVY), align=A_LEFT)
cell(ws, "F11", "Total miles (year)", font=F_TH, fillc=fill(NAVY), align=A_LEFT)

vehicles = [
    ("Main car",      "2020 Toyota Corolla", 38500, 78500),
    ("Second car",    "",                    "",    ""),
    ("Backup / rental","",                   "",    ""),
]
for i, (name, ymm, sod, eod) in enumerate(vehicles):
    r = 12 + i
    ws.row_dimensions[r].height = 22
    cell(ws, f"B{r}", name, font=F_BODY, fillc=fill(GOLD_PALE), align=A_LEFT, border=thin_box(GOLD))
    cell(ws, f"C{r}", ymm,  font=F_BODY, fillc=fill(GOLD_PALE), align=A_LEFT, border=thin_box(GOLD))
    cell(ws, f"D{r}", sod,  font=F_BODY, fillc=fill(GOLD_PALE), align=A_RIGHT, border=thin_box(GOLD), fmt='#,##0')
    cell(ws, f"E{r}", eod,  font=F_BODY, fillc=fill(GOLD_PALE), align=A_RIGHT, border=thin_box(GOLD), fmt='#,##0')
    # Calculated column
    f = ws[f"F{r}"]
    f.value = f'=IF(AND(ISNUMBER(D{r}),ISNUMBER(E{r})),E{r}-D{r},"")'
    f.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    f.fill = fill(CREAM_D)
    f.alignment = A_RIGHT
    f.border = thin_box(BORDER)
    f.number_format = '#,##0'

# Named range for vehicle dropdown (only names, B12:B14)
wb.defined_names["VehicleList"] = DefinedName("VehicleList", attr_text="Setup!$B$12:$B$14")

# Your platforms
ws.row_dimensions[16].height = 24
cell(ws, "B16", "Your gig platforms", font=F_H1, align=A_LEFT)
ws.merge_cells("B16:F16")
cell(ws, "B17", "Edit this list to match the apps you actually drive for.", font=F_MUTED, align=A_LEFT)
ws.merge_cells("B17:F17")

platforms = ["DoorDash", "Uber", "Uber Eats", "Lyft", "Grubhub", "Instacart",
             "Amazon Flex", "TaskRabbit", "Rover", "Shipt", "Roadie", "Other"]
ws.row_dimensions[18].height = 22
cell(ws, "B18", "Platform", font=F_TH, fillc=fill(NAVY), align=A_LEFT)
for i, p in enumerate(platforms):
    r = 19 + i
    cell(ws, f"B{r}", p, font=F_BODY, fillc=fill(GOLD_PALE), align=A_LEFT, border=thin_box(GOLD))
    ws.row_dimensions[r].height = 20

wb.defined_names["PlatformList"] = DefinedName(
    "PlatformList", attr_text=f"Setup!$B$19:$B${18+len(platforms)}"
)

# ════════════════════════════════════════════════════════════════════
# SHEET 3 — TRIP LOG
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Trip Log")
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110
ws.freeze_panes = "A3"

ws.column_dimensions["A"].width = 12   # Date
ws.column_dimensions["B"].width = 18   # Vehicle
ws.column_dimensions["C"].width = 16   # Platform
ws.column_dimensions["D"].width = 28   # Purpose / route
ws.column_dimensions["E"].width = 12   # Start odom
ws.column_dimensions["F"].width = 12   # End odom
ws.column_dimensions["G"].width = 11   # Miles (manual or calc)
ws.column_dimensions["H"].width = 11   # Business %
ws.column_dimensions["I"].width = 13   # Business miles
ws.column_dimensions["J"].width = 13   # Deduction $
ws.column_dimensions["K"].width = 22   # Notes

ws.row_dimensions[1].height = 30
cell(ws, "A1", "Trip Log — one row per business trip. Use dropdowns where they appear.",
     font=Font(name="Calibri", size=11, italic=True, color=GRAY), align=A_LEFT)
ws.merge_cells("A1:K1")

# Header row
headers = ["Date","Vehicle","Platform","Purpose / Route",
           "Start odom.","End odom.","Miles","Business %","Business miles","Deduction $","Notes"]
ws.row_dimensions[2].height = 32
for i, h in enumerate(headers):
    c = ws.cell(row=2, column=i+1, value=h)
    c.font = F_TH
    c.fill = fill(NAVY)
    c.alignment = A_CTR
    c.border = thin_box(NAVY)

# Number of data rows
DATA_ROWS = 365  # one per day, room for the year

# Pre-populated example rows (first 2 rows show users how to log)
examples = [
    ["2026-01-03", "Main car", "DoorDash", "Long Beach lunch shift",   38500, 38582, None, 1.00, "First day driving — see formulas below for how miles/deduction auto-fill"],
    ["2026-01-04", "Main car", "Uber",     "Friday airport runs",       None,  None,  73, 1.00, "Used miles directly instead of odometer — both work"],
]
for i, ex in enumerate(examples):
    r = 3 + i
    ws.row_dimensions[r].height = 22
    ws.cell(row=r, column=1, value=ex[0]).number_format = "yyyy-mm-dd"
    ws.cell(row=r, column=2, value=ex[1])
    ws.cell(row=r, column=3, value=ex[2])
    ws.cell(row=r, column=4, value=ex[3])
    if ex[4] is not None: ws.cell(row=r, column=5, value=ex[4]).number_format = "#,##0"
    if ex[5] is not None: ws.cell(row=r, column=6, value=ex[5]).number_format = "#,##0"
    if ex[6] is not None: ws.cell(row=r, column=7, value=ex[6])
    ws.cell(row=r, column=8, value=ex[7]).number_format = "0%"
    ws.cell(row=r, column=11, value=ex[8])
    # Light gold tint to show "example"
    for col in range(1, 12):
        c = ws.cell(row=r, column=col)
        c.alignment = A_LEFT if col != 8 else A_CTR
        c.fill = fill(GOLD_PALE)
        c.border = thin_box(BORDER)
        if c.font.size is None:
            c.font = Font(name="Calibri", size=10, italic=True, color=NAVY)
        else:
            c.font = Font(name="Calibri", size=10, italic=True, color=NAVY)

# Add formulas for example rows
for r in (3, 4):
    # Column G — Miles: if user entered odom both, compute; otherwise leave manual
    # We pre-filled row 3 with odometers (let formula compute); row 4 with miles directly
    if r == 3:
        ws.cell(row=r, column=7).value = f"=IF(AND(ISNUMBER(E{r}),ISNUMBER(F{r})),F{r}-E{r},\"\")"
        ws.cell(row=r, column=7).number_format = "#,##0"
    # Column I — Business miles
    ws.cell(row=r, column=9).value = f"=IF(AND(ISNUMBER(G{r}),ISNUMBER(H{r})),G{r}*H{r},\"\")"
    ws.cell(row=r, column=9).number_format = "#,##0.0"
    # Column J — Deduction
    ws.cell(row=r, column=10).value = f"=IF(ISNUMBER(I{r}),I{r}*MileRate,\"\")"
    ws.cell(row=r, column=10).number_format = '"$"#,##0.00'

# Blank rows for user entries — pre-load formulas in G, I, J + default 100% business
for r in range(5, 5 + DATA_ROWS):
    ws.row_dimensions[r].height = 20
    # Date column blank, formatted
    ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
    # Numeric formats
    ws.cell(row=r, column=5).number_format = "#,##0"
    ws.cell(row=r, column=6).number_format = "#,##0"
    # G — Miles: shows odometer subtraction if both filled, else stays manual-editable
    ws.cell(row=r, column=7).value = f'=IF(AND(ISNUMBER(E{r}),ISNUMBER(F{r})),F{r}-E{r},"")'
    ws.cell(row=r, column=7).number_format = "#,##0"
    # H — default Business %
    ws.cell(row=r, column=8).value = 1.0
    ws.cell(row=r, column=8).number_format = "0%"
    ws.cell(row=r, column=8).alignment = A_CTR
    # I — Business miles
    ws.cell(row=r, column=9).value = f'=IF(AND(ISNUMBER(G{r}),ISNUMBER(H{r})),G{r}*H{r},"")'
    ws.cell(row=r, column=9).number_format = "#,##0.0"
    # J — Deduction $
    ws.cell(row=r, column=10).value = f'=IF(ISNUMBER(I{r}),I{r}*MileRate,"")'
    ws.cell(row=r, column=10).number_format = '"$"#,##0.00'
    # Borders for clean look
    for col in range(1, 12):
        ws.cell(row=r, column=col).border = thin_box(BORDER)
        ws.cell(row=r, column=col).font = F_BODY

# Data validation: Vehicle dropdown
dv_vehicle = DataValidation(type="list", formula1="=VehicleList", allow_blank=True)
dv_vehicle.error = "Pick a vehicle from the dropdown, or add it on the Setup tab."
dv_vehicle.errorTitle = "Unknown vehicle"
ws.add_data_validation(dv_vehicle)
dv_vehicle.add(f"B3:B{4+DATA_ROWS}")

# Data validation: Platform dropdown
dv_platform = DataValidation(type="list", formula1="=PlatformList", allow_blank=True)
dv_platform.error = "Pick a platform from the dropdown, or add it on the Setup tab."
dv_platform.errorTitle = "Unknown platform"
ws.add_data_validation(dv_platform)
dv_platform.add(f"C3:C{4+DATA_ROWS}")

# Business % validation: 0% – 100%
dv_pct = DataValidation(type="decimal", operator="between", formula1=0, formula2=1, allow_blank=True)
dv_pct.error = "Business % must be between 0% and 100%."
dv_pct.errorTitle = "Invalid business %"
ws.add_data_validation(dv_pct)
dv_pct.add(f"H3:H{4+DATA_ROWS}")

# Set up an Excel Table for the log so it stays styled as users add rows
# (Note: openpyxl tables can have rough edges; we keep formulas + formatting baked into each row instead.)

# ════════════════════════════════════════════════════════════════════
# SHEET 4 — MONTHLY SUMMARY
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Monthly Summary")
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 16

ws.row_dimensions[2].height = 32
cell(ws, "B2", "Monthly Summary", font=Font(name="Calibri", size=22, bold=True, color=NAVY), align=A_LEFT)
ws.merge_cells("B2:F2")
cell(ws, "B3", "Pulled from the Trip Log automatically. Edit nothing here.", font=F_MUTED, align=A_LEFT)
ws.merge_cells("B3:F3")

# Header
ws.row_dimensions[5].height = 26
for i, h in enumerate(["Month","Trips","Total miles","Business miles","Deduction $"]):
    c = ws.cell(row=5, column=2+i, value=h)
    c.font = F_TH
    c.fill = fill(NAVY)
    c.alignment = A_CTR
    c.border = thin_box(NAVY)

months = ["January","February","March","April","May","June",
          "July","August","September","October","November","December"]

LOG = "'Trip Log'"
# Row range to look at on Trip Log (rows 3..end of data)
log_range_date = f"{LOG}!$A$3:$A$369"
log_range_miles = f"{LOG}!$G$3:$G$369"
log_range_busmi = f"{LOG}!$I$3:$I$369"
log_range_ded   = f"{LOG}!$J$3:$J$369"

for i, m in enumerate(months):
    r = 6 + i
    ws.row_dimensions[r].height = 22
    cell(ws, f"B{r}", m, font=F_BODY, align=A_LEFT, border=thin_box(BORDER))
    # Trips: count of dates in this month/year
    ws.cell(row=r, column=3).value = (
        f'=SUMPRODUCT((MONTH({log_range_date})={i+1})*(YEAR({log_range_date})=TaxYear)*ISNUMBER({log_range_miles}))'
    )
    # Total miles
    ws.cell(row=r, column=4).value = (
        f'=SUMPRODUCT((MONTH({log_range_date})={i+1})*(YEAR({log_range_date})=TaxYear)*IFERROR({log_range_miles},0))'
    )
    # Business miles
    ws.cell(row=r, column=5).value = (
        f'=SUMPRODUCT((MONTH({log_range_date})={i+1})*(YEAR({log_range_date})=TaxYear)*IFERROR({log_range_busmi},0))'
    )
    # Deduction $
    ws.cell(row=r, column=6).value = (
        f'=SUMPRODUCT((MONTH({log_range_date})={i+1})*(YEAR({log_range_date})=TaxYear)*IFERROR({log_range_ded},0))'
    )
    # Formats
    ws.cell(row=r, column=3).number_format = "#,##0"
    ws.cell(row=r, column=4).number_format = "#,##0.0"
    ws.cell(row=r, column=5).number_format = "#,##0.0"
    ws.cell(row=r, column=6).number_format = '"$"#,##0.00'
    for c in range(3,7):
        ws.cell(row=r, column=c).alignment = A_RIGHT
        ws.cell(row=r, column=c).border = thin_box(BORDER)

# Totals row
totals_row = 18
ws.row_dimensions[totals_row].height = 28
cell(ws, f"B{totals_row}", "Total", font=F_TOTAL, fillc=fill(GOLD_PALE), align=A_LEFT, border=thin_box(GOLD))
for col_letter, idx in zip(["C","D","E","F"], [3,4,5,6]):
    c = ws.cell(row=totals_row, column=idx)
    c.value = f"=SUM({col_letter}6:{col_letter}17)"
    c.font = F_TOTAL
    c.fill = fill(GOLD_PALE)
    c.alignment = A_RIGHT
    c.border = thin_box(GOLD)
    c.number_format = '#,##0.0' if col_letter in ("D","E") else ('#,##0' if col_letter=="C" else '"$"#,##0.00')

# Platform breakdown
ws.row_dimensions[21].height = 26
cell(ws, "B21", "By platform (full year)", font=F_H1, align=A_LEFT)
ws.merge_cells(f"B21:F21")
ws.row_dimensions[22].height = 24
for i, h in enumerate(["Platform","Trips","Business miles","Deduction $"]):
    c = ws.cell(row=22, column=2+i, value=h)
    c.font = F_TH
    c.fill = fill(NAVY)
    c.alignment = A_CTR
    c.border = thin_box(NAVY)

# Pull from PlatformList — rows 19..30 on Setup
for i in range(12):  # max 12 platforms
    r = 23 + i
    setup_r = 19 + i
    ws.row_dimensions[r].height = 20
    cell(ws, f"B{r}", f"=IFERROR(Setup!B{setup_r},\"\")", font=F_BODY, align=A_LEFT, border=thin_box(BORDER))
    # Trips
    ws.cell(row=r, column=3).value = (
        f'=IF(B{r}="","",SUMPRODUCT(({LOG}!$C$3:$C$369=B{r})*(YEAR({log_range_date})=TaxYear)*ISNUMBER({log_range_miles})))'
    )
    ws.cell(row=r, column=4).value = (
        f'=IF(B{r}="","",SUMPRODUCT(({LOG}!$C$3:$C$369=B{r})*(YEAR({log_range_date})=TaxYear)*IFERROR({log_range_busmi},0)))'
    )
    ws.cell(row=r, column=5).value = (
        f'=IF(B{r}="","",SUMPRODUCT(({LOG}!$C$3:$C$369=B{r})*(YEAR({log_range_date})=TaxYear)*IFERROR({log_range_ded},0)))'
    )
    ws.cell(row=r, column=3).number_format = "#,##0"
    ws.cell(row=r, column=4).number_format = "#,##0.0"
    ws.cell(row=r, column=5).number_format = '"$"#,##0.00'
    for c in range(3, 6):
        ws.cell(row=r, column=c).alignment = A_RIGHT
        ws.cell(row=r, column=c).border = thin_box(BORDER)

# ════════════════════════════════════════════════════════════════════
# SHEET 5 — YEAR-END SUMMARY
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Year-End Summary")
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110

for col, w in zip("ABCDEFGH", [2, 24, 24, 24, 24, 2, 2, 2]):
    ws.column_dimensions[col].width = w

ws.row_dimensions[2].height = 32
cell(ws, "B2", "Year-End Summary", font=Font(name="Calibri", size=22, bold=True, color=NAVY), align=A_LEFT)
ws.merge_cells("B2:E2")
cell(ws, "B3", "What to put on Schedule C line 9 (Car & Truck Expenses) when you file.", font=F_MUTED, align=A_LEFT)
ws.merge_cells("B3:E3")

# KPI cards
def kpi(top_row, col_letter, label, value_formula, fmt, color=NAVY):
    # Label
    cl = ws[f"{col_letter}{top_row}"]
    cl.value = label
    cl.font = F_KPI_LBL
    cl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cl.fill = fill(CREAM)
    cl.border = Border(
        left=Side(style="thin", color=BORDER),
        right=Side(style="thin", color=BORDER),
        top=Side(style="thin", color=BORDER),
        bottom=Side(style=None)
    )
    ws.row_dimensions[top_row].height = 24
    # Value
    cv = ws[f"{col_letter}{top_row+1}"]
    cv.value = value_formula
    cv.font = Font(name="Calibri", size=22, bold=True, color=color)
    cv.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cv.fill = fill(CREAM)
    cv.number_format = fmt
    cv.border = Border(
        left=Side(style="thin", color=BORDER),
        right=Side(style="thin", color=BORDER),
        top=Side(style=None),
        bottom=Side(style="thin", color=BORDER)
    )
    ws.row_dimensions[top_row+1].height = 36

kpi(5, "B", "TAX YEAR",        "=TaxYear",                       "0",         NAVY)
kpi(5, "C", "TOTAL BUSINESS MILES", "='Monthly Summary'!E18",    "#,##0.0",   NAVY)
kpi(5, "D", "IRS RATE",        "=MileRate",                       '"$"#,##0.000', NAVY)
kpi(5, "E", "MILEAGE DEDUCTION","='Monthly Summary'!F18",        '"$"#,##0.00', GOLD.replace("FF","FF"))

# Schedule C callout
ws.row_dimensions[9].height = 10
ws.row_dimensions[10].height = 24
cell(ws, "B10", "Schedule C — Part II, Line 9 (Car & Truck Expenses):", font=F_H2, align=A_LEFT)
ws.merge_cells("B10:E10")

cell(ws, "B11", "=\"Enter $\"&TEXT('Monthly Summary'!F18,\"#,##0.00\")&\" on line 9. Use the 'Standard Mileage' method (the option you've been tracking for).\"",
     font=F_BODY, align=A_LEFT)
ws.merge_cells("B11:E11")
ws.row_dimensions[11].height = 28

# Reasonableness check
ws.row_dimensions[14].height = 24
cell(ws, "B14", "Sanity check vs. your total driving", font=F_H1, align=A_LEFT)
ws.merge_cells("B14:E14")

cell(ws, "B15", "Total miles all vehicles drove this year (from Setup):", font=F_BODY, align=A_LEFT)
ws["C15"].value = "=SUM(Setup!F12:F14)"
ws["C15"].number_format = "#,##0"
ws["C15"].font = F_TOTAL
ws["C15"].alignment = A_RIGHT

cell(ws, "B16", "Business miles you logged this year:", font=F_BODY, align=A_LEFT)
ws["C16"].value = "='Monthly Summary'!E18"
ws["C16"].number_format = "#,##0.0"
ws["C16"].font = F_TOTAL
ws["C16"].alignment = A_RIGHT

cell(ws, "B17", "Implied business-use percentage:", font=F_BODY, align=A_LEFT)
ws["C17"].value = "=IFERROR(C16/C15,\"\")"
ws["C17"].number_format = "0.0%"
ws["C17"].font = Font(name="Calibri", size=12, bold=True, color=NAVY)
ws["C17"].alignment = A_RIGHT

# Conditional formatting on C17: green if <80%, amber if 80-95%, red >95%
ws.conditional_formatting.add("C17",
    CellIsRule(operator="lessThan",    formula=["0.8"], fill=fill(GREEN_BG), font=Font(color=GREEN, bold=True)))
ws.conditional_formatting.add("C17",
    CellIsRule(operator="between",     formula=["0.8","0.95"], fill=fill(AMBER_BG), font=Font(color=AMBER, bold=True)))
ws.conditional_formatting.add("C17",
    CellIsRule(operator="greaterThan", formula=["0.95"], fill=fill("FFFBE9E9"), font=Font(color="FFB83232", bold=True)))

ws.row_dimensions[18].height = 50
cell(ws, "B18",
     "If this percentage is above ~85%, expect IRS questions — most gig drivers fall in the 50–85% range because of commute, personal errands, and time between gigs. Reasonable numbers help you survive an audit.",
     font=F_MUTED, align=A_LEFT_T)
ws.merge_cells("B18:E18")

# Records to keep
ws.row_dimensions[21].height = 24
cell(ws, "B21", "Records to keep with this file (3+ years)", font=F_H1, align=A_LEFT)
ws.merge_cells("B21:E21")

records = [
    "Jan 1 and Dec 31 odometer photos for every vehicle you used",
    "1099-NEC and 1099-K from every platform",
    "Bank/PayPal statements showing platform deposits",
    "Receipts for tolls, parking, and any non-mileage car costs (these add on top of the mileage deduction)",
    "If you ever switched from standard mileage to actual expenses, keep that election documented",
]
for i, item in enumerate(records):
    r = 22 + i
    ws.row_dimensions[r].height = 22
    cell(ws, f"B{r}", "→", font=Font(name="Calibri", size=12, bold=True, color=GOLD), align=A_CTR)
    cell(ws, f"C{r}", item, font=F_BODY, align=A_LEFT)
    ws.merge_cells(f"C{r}:E{r}")

# ════════════════════════════════════════════════════════════════════
# SHEET 6 — IRS REFERENCE
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("IRS Reference")
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 70

ws.row_dimensions[2].height = 32
cell(ws, "B2", "Plain-English IRS Reference", font=Font(name="Calibri", size=22, bold=True, color=NAVY), align=A_LEFT)
ws.merge_cells("B2:C2")
cell(ws, "B3", "What every gig driver needs to know — short answers to the questions the IRS actually cares about.", font=F_MUTED, align=A_LEFT)
ws.merge_cells("B3:C3")

faqs = [
    ("Standard mileage vs. actual expenses",
     "You pick one method per vehicle. Standard mileage (this tracker) multiplies your business miles by the IRS rate — no receipts needed beyond the log. Actual expenses adds up gas, oil, repairs, insurance, depreciation, etc., then multiplies by business-use %. Most gig drivers come out ahead with standard mileage and it's way less paperwork."),
    ("Can I switch methods later?",
     "If you used STANDARD mileage in year one, you can switch to actual in any later year. If you used ACTUAL in year one, you're locked into actual for that vehicle for as long as you own it. This is why year-one choice matters."),
    ("What counts as a business mile?",
     "Driving with the app on and waiting for/accepting/completing rides or deliveries. Driving from one delivery to the next while online. Driving home after your shift IS a business mile if you're online and accepting requests. Plain personal errands are not."),
    ("Is my commute deductible?",
     "Your drive from home to your first business pickup is generally NOT deductible (commuting). Once you're logged into the app and accepting work, you've started your business — miles from that point until you log off ARE deductible. Some drivers turn the app on as they leave the driveway specifically for this reason."),
    ("Do tolls and parking count?",
     "Yes, on top of the mileage deduction. Track them separately and add them to Schedule C — they don't go in this tracker because they aren't mileage."),
    ("What about my phone, car wash, and hot bag?",
     "Phone (business-use %), car washes, delivery bags, accessories, and required gear go on Schedule C as separate expenses — not in this tracker. The mileage deduction only covers what the IRS calls the operating cost of the vehicle."),
    ("How long do I keep this log?",
     "At least 3 years after you file. The IRS can audit up to 3 years routinely, 6 years for substantial under-reporting, and indefinitely for fraud. Save a PDF export each year and stash it with your tax records."),
    ("What if I drove for two platforms at the same time?",
     "Log it once. You can't double-count miles. Pick the platform you were primarily working for on that trip."),
    ("Do I need a paper log too?",
     "No. The IRS accepts digital records as long as they're contemporaneous (entered close to when the trip happened) and complete. This spreadsheet meets that standard."),
    ("What if I forget to log a day?",
     "Reconstruct it from your platform's trip history (DoorDash, Uber, Lyft, etc. all keep one) within a few days. The IRS calls this 'reasonable reconstruction.' Don't wait until April."),
]

row = 5
for q, a in faqs:
    ws.row_dimensions[row].height = 22
    cell(ws, f"B{row}", "Q", font=Font(name="Calibri", size=12, bold=True, color=GOLD), align=A_CTR)
    cell(ws, f"C{row}", q, font=F_H2, align=A_LEFT)
    row += 1
    ws.row_dimensions[row].height = 60
    cell(ws, f"C{row}", a, font=F_BODY, align=A_LEFT_T)
    row += 2  # blank spacer

# ════════════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════════════
import os
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "gig-driver-mileage-tracker-2026.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Size: {os.path.getsize(out_path)/1024:.1f} KB")
