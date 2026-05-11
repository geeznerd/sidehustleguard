"""
Build the Quarterly Tax System 2026 — Excel/Google Sheets product.
Outputs: quarterly-tax-system-2026.xlsx in the same directory.
"""
import openpyxl, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# ── COLORS ──
NAVY      = "FF1C2B4A"
GOLD      = "FFC9973A"
GOLD_LT   = "FFDDB06A"
GOLD_PALE = "FFFDF5E8"
CREAM     = "FFFAF8F4"
CREAM_D   = "FFF0ECE3"
GREEN     = "FF276944"
GREEN_BG  = "FFEAF5EF"
AMBER     = "FFA86C0A"
AMBER_BG  = "FFFEF6E8"
RED       = "FFB83232"
RED_BG    = "FFFBE9E9"
WHITE     = "FFFFFFFF"
GRAY      = "FF6B7A96"
BORDER    = "FFE5E7ED"

def fill(c): return PatternFill(start_color=c, end_color=c, fill_type="solid")
def box(c=BORDER):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def cell(ws, coord, value, *, font=None, fillc=None, align=None, border=None, fmt=None):
    c = ws[coord]
    c.value = value
    if font: c.font = font
    if fillc: c.fill = fillc
    if align: c.alignment = align
    if border: c.border = border
    if fmt: c.number_format = fmt
    return c

F_TITLE    = Font(name="Calibri", size=26, bold=True, color=NAVY)
F_SUBTITLE = Font(name="Calibri", size=11, italic=True, color=GOLD)
F_H1       = Font(name="Calibri", size=15, bold=True, color=NAVY)
F_H2       = Font(name="Calibri", size=12, bold=True, color=NAVY)
F_BODY     = Font(name="Calibri", size=11, color="FF333333")
F_MUTED    = Font(name="Calibri", size=10, color=GRAY)
F_TH       = Font(name="Calibri", size=11, bold=True, color=WHITE)
F_TOTAL    = Font(name="Calibri", size=12, bold=True, color=NAVY)
F_KPI_LBL  = Font(name="Calibri", size=10, bold=True, color=GRAY)
F_KPI_VAL  = Font(name="Calibri", size=22, bold=True, color=NAVY)
F_INPUT    = Font(name="Calibri", size=12, bold=True, color=NAVY)
F_CALC     = Font(name="Calibri", size=11, bold=True, color=NAVY)

A_LEFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
A_LEFT_T= Alignment(horizontal="left",   vertical="top",    wrap_text=True)
A_CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_RIGHT = Alignment(horizontal="right",  vertical="center")

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════
# SHEET 1 — START HERE
# ════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Start Here"
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110

ws.column_dimensions["A"].width = 2
for col, w in zip("BCDEFGH", [22, 18, 18, 18, 18, 18, 18]):
    ws.column_dimensions[col].width = w

ws.row_dimensions[2].height = 38
cell(ws, "B2", "Quarterly Tax System 2026", font=F_TITLE, align=A_LEFT)
ws.merge_cells("B2:H2")
cell(ws, "B3", "SideHustleGuard · sidehustleguard.com", font=F_SUBTITLE, align=A_LEFT)
ws.merge_cells("B3:H3")

ws.row_dimensions[5].height = 6
ws.row_dimensions[6].height = 22
cell(ws, "B6", "What this tool does", font=F_H1, align=A_LEFT)
ws.merge_cells("B6:H6")

intro = ("Tells you exactly how much to send the IRS (and your state) each quarter so you never get hit with "
         "an underpayment penalty in April. Works for any self-employed income — gig driving, Airbnb, Etsy, "
         "freelance, consulting, multiple streams. Builds a complete record of every payment so when you sit "
         "down with your CPA in March, the work is already done.")
ws.row_dimensions[7].height = 85
cell(ws, "B7", intro, font=F_BODY, align=A_LEFT_T)
ws.merge_cells("B7:H7")

ws.row_dimensions[9].height = 22
cell(ws, "B9", "How to use it (3 steps)", font=F_H1, align=A_LEFT)
ws.merge_cells("B9:H9")

steps = [
    ("1.", "Fill in the Setup tab.",
     "Filing status, state, what you expect to make this year, what you owed in tax last year, and any W-2 income from a day job. Defaults are pre-filled — just edit the gold cells."),
    ("2.", "Forecast your year on the Income Forecast tab.",
     "Drop in your expected net income (after deductions like mileage) for each month. You can break it down by platform (DoorDash + Etsy + freelance) or just enter one annual number."),
    ("3.", "Check the Quarterly Payments tab before each due date.",
     "It shows you exactly what to send the IRS for Q1, Q2, Q3, Q4, and the state. Log each payment in the same tab so you have a complete record."),
]
row = 10
for num, head, body in steps:
    ws.row_dimensions[row].height = 56
    cell(ws, "B"+str(row), num, font=Font(name="Calibri", size=22, bold=True, color=GOLD), align=Alignment(horizontal="center", vertical="top"))
    cell(ws, "C"+str(row), head, font=F_H2, align=A_LEFT_T)
    cell(ws, "D"+str(row), body, font=F_BODY, align=A_LEFT_T)
    ws.merge_cells("D{0}:H{0}".format(row))
    row += 1

# Due dates table
row += 1
ws.row_dimensions[row].height = 22
cell(ws, "B"+str(row), "Quarterly due dates", font=F_H1, align=A_LEFT)
ws.merge_cells("B{0}:H{0}".format(row))
row += 1
ws.row_dimensions[row].height = 24
cell(ws, "B"+str(row), "Quarter", font=F_TH, fillc=fill(NAVY), align=A_CTR)
cell(ws, "C"+str(row), "Covers income from", font=F_TH, fillc=fill(NAVY), align=A_CTR)
cell(ws, "D"+str(row), "Federal due", font=F_TH, fillc=fill(NAVY), align=A_CTR)
ws.merge_cells(f"D{row}:E{row}")
cell(ws, "F"+str(row), "State due (most states)", font=F_TH, fillc=fill(NAVY), align=A_CTR)
ws.merge_cells(f"F{row}:H{row}")
row += 1

quarters = [
    ("Q1", "Jan – Mar", "April 15", "April 15 (same)"),
    ("Q2", "Apr – May", "June 15", "June 15 (same)"),
    ("Q3", "Jun – Aug", "September 15", "September 15 (same)"),
    ("Q4", "Sep – Dec", "January 15 (next year)", "January 15 (next year)"),
]
for q, span, fed, state in quarters:
    ws.row_dimensions[row].height = 22
    cell(ws, f"B{row}", q,     font=F_BODY, align=A_CTR,   border=box())
    cell(ws, f"C{row}", span,  font=F_BODY, align=A_LEFT,  border=box())
    cell(ws, f"D{row}", fed,   font=F_CALC, align=A_LEFT,  border=box())
    ws.merge_cells(f"D{row}:E{row}")
    cell(ws, f"F{row}", state, font=F_BODY, align=A_LEFT,  border=box())
    ws.merge_cells(f"F{row}:H{row}")
    row += 1

# Penalty callout
row += 1
ws.row_dimensions[row].height = 22
cell(ws, "B"+str(row), "Why this matters", font=F_H1, align=A_LEFT)
ws.merge_cells("B{0}:H{0}".format(row))
row += 1
ws.row_dimensions[row].height = 80
penalty = ("If you owe more than $1,000 in tax and didn't make quarterly payments — or paid less than the "
           "\"safe harbor\" amount — the IRS adds an underpayment penalty on top of what you owe. The penalty "
           "is calculated quarterly at the federal short-term rate plus 3% (typically 6–8% annualized). On a "
           "$15,000 tax bill, that's an extra $400–$800 that didn't have to exist. This tool exists to make "
           "sure you hit one of the two safe-harbor tests so the penalty never applies.")
cell(ws, "B"+str(row), penalty, font=F_BODY, align=A_LEFT_T)
ws.merge_cells("B{0}:H{0}".format(row))

# Disclaimer
row += 2
ws.row_dimensions[row].height = 50
disc = ("Educational tool only — not tax or legal advice. Tax brackets, SE tax wage base, and state rates change "
        "annually. Verify all defaults against current IRS and state publications. For complex situations (W-2 + SE income, "
        "joint filers with separate businesses, large investment income), consult a CPA.")
cell(ws, "B"+str(row), disc, font=Font(name="Calibri", size=9, italic=True, color=GRAY), align=A_LEFT_T)
ws.merge_cells("B{0}:H{0}".format(row))

# ════════════════════════════════════════════════════════════════════
# SHEET 2 — SETUP
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Setup")
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 50

ws.row_dimensions[2].height = 32
cell(ws, "B2", "Setup", font=Font(name="Calibri", size=22, bold=True, color=NAVY), align=A_LEFT)
ws.merge_cells("B2:D2")
cell(ws, "B3", "Edit the gold cells. Everything else reads from here.", font=F_MUTED, align=A_LEFT)
ws.merge_cells("B3:D3")

# Helper to lay out one input row
def input_row(r, label, value, hint, fmt=None, is_input=True):
    cell(ws, f"B{r}", label, font=F_BODY, align=A_LEFT)
    c = ws[f"C{r}"]
    c.value = value
    c.font = F_INPUT
    c.alignment = A_RIGHT
    c.fill = fill(GOLD_PALE) if is_input else fill(CREAM_D)
    c.border = box(GOLD if is_input else BORDER)
    if fmt: c.number_format = fmt
    cell(ws, f"D{r}", hint, font=F_MUTED, align=A_LEFT)
    ws.row_dimensions[r].height = 24

# ── Section A: Year & filing
row = 5
ws.row_dimensions[row].height = 24
cell(ws, f"B{row}", "Year & filing status", font=F_H1, align=A_LEFT)
ws.merge_cells(f"B{row}:D{row}")
row += 1
input_row(row, "Tax year", 2026, "The year you're paying taxes for")
TAX_YEAR_CELL = f"Setup!$C${row}"
row += 1
input_row(row, "Filing status", "Single", "Single, MFJ, MFS, HOH")
FILING_CELL = f"Setup!$C${row}"
# Validation
dv = DataValidation(type="list", formula1='"Single,MFJ,MFS,HOH"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(f"C{row}")
row += 1
input_row(row, "State", "CA", "Two-letter code (e.g., CA, TX, FL, NY)")
STATE_CELL = f"Setup!$C${row}"
row += 1
input_row(row, "State flat estimated tax rate", 0.06, "Override the lookup below if you know your effective rate. Edit at any time.", fmt="0.0%")
STATE_RATE_CELL = f"Setup!$C${row}"

# ── Section B: This year's income
row += 2
ws.row_dimensions[row].height = 24
cell(ws, f"B{row}", "This year's income (overrides if filled)", font=F_H1, align=A_LEFT)
ws.merge_cells(f"B{row}:D{row}")
row += 1
input_row(row, "Estimated net self-employment income (annual)", 0, "Leave 0 to use the month-by-month forecast on the Income Forecast tab", fmt='"$"#,##0')
SE_OVERRIDE_CELL = f"Setup!$C${row}"
row += 1
input_row(row, "W-2 income (if you also have a day job)", 0, "Gross wages, before tax", fmt='"$"#,##0')
W2_INCOME_CELL = f"Setup!$C${row}"
row += 1
input_row(row, "Federal tax withheld from W-2 (this year)", 0, "From your most recent paystub × pay periods left, or your W-2 box 2", fmt='"$"#,##0')
W2_WITHHELD_CELL = f"Setup!$C${row}"

# ── Section C: Last year (for safe harbor)
row += 2
ws.row_dimensions[row].height = 24
cell(ws, f"B{row}", "Last year — for safe harbor calculation", font=F_H1, align=A_LEFT)
ws.merge_cells(f"B{row}:D{row}")
row += 1
input_row(row, "Total federal tax owed last year (Form 1040 line 24)", 0, "Look at last year's return. This sets the 100%/110% safe harbor.", fmt='"$"#,##0')
LAST_TAX_CELL = f"Setup!$C${row}"
row += 1
input_row(row, "Last year's AGI (Form 1040 line 11)", 0, "If above $150K, safe harbor is 110% of last year's tax instead of 100%", fmt='"$"#,##0')
LAST_AGI_CELL = f"Setup!$C${row}"

# ── Section D: Federal rates (editable for any tax year)
row += 2
ws.row_dimensions[row].height = 24
cell(ws, f"B{row}", "Federal tax constants (edit each year)", font=F_H1, align=A_LEFT)
ws.merge_cells(f"B{row}:D{row}")
row += 1
input_row(row, "Standard deduction (single / MFS)", 15750, "2025 value as published by IRS. Update for your tax year.", fmt='"$"#,##0')
STD_DED_SINGLE = f"Setup!$C${row}"
row += 1
input_row(row, "Standard deduction (MFJ)", 31500, "2025 value", fmt='"$"#,##0')
STD_DED_MFJ = f"Setup!$C${row}"
row += 1
input_row(row, "Standard deduction (HOH)", 23625, "2025 value", fmt='"$"#,##0')
STD_DED_HOH = f"Setup!$C${row}"
row += 1
input_row(row, "Social Security wage base", 176100, "2025 value. SE tax SS portion (12.4%) only applies up to this.", fmt='"$"#,##0')
SS_BASE_CELL = f"Setup!$C${row}"

# Named ranges to make formulas readable
wb.defined_names["TaxYear"]      = DefinedName("TaxYear",      attr_text=TAX_YEAR_CELL)
wb.defined_names["FilingStatus"] = DefinedName("FilingStatus", attr_text=FILING_CELL)
wb.defined_names["StateCode"]    = DefinedName("StateCode",    attr_text=STATE_CELL)
wb.defined_names["StateRate"]    = DefinedName("StateRate",    attr_text=STATE_RATE_CELL)
wb.defined_names["SEOverride"]   = DefinedName("SEOverride",   attr_text=SE_OVERRIDE_CELL)
wb.defined_names["W2Income"]     = DefinedName("W2Income",     attr_text=W2_INCOME_CELL)
wb.defined_names["W2Withheld"]   = DefinedName("W2Withheld",   attr_text=W2_WITHHELD_CELL)
wb.defined_names["LastYearTax"]  = DefinedName("LastYearTax",  attr_text=LAST_TAX_CELL)
wb.defined_names["LastYearAGI"]  = DefinedName("LastYearAGI",  attr_text=LAST_AGI_CELL)
wb.defined_names["StdDedSingle"] = DefinedName("StdDedSingle", attr_text=STD_DED_SINGLE)
wb.defined_names["StdDedMFJ"]    = DefinedName("StdDedMFJ",    attr_text=STD_DED_MFJ)
wb.defined_names["StdDedHOH"]    = DefinedName("StdDedHOH",    attr_text=STD_DED_HOH)
wb.defined_names["SSWageBase"]   = DefinedName("SSWageBase",   attr_text=SS_BASE_CELL)

# ── Section E: Federal brackets (single + MFJ + HOH)
row += 2
ws.row_dimensions[row].height = 24
cell(ws, f"B{row}", "Federal income tax brackets (2025 — edit each year)", font=F_H1, align=A_LEFT)
ws.merge_cells(f"B{row}:D{row}")
row += 1
ws.row_dimensions[row].height = 22
cell(ws, f"B{row}", "Bracket top (Single)", font=F_TH, fillc=fill(NAVY), align=A_CTR)
cell(ws, f"C{row}", "Bracket top (MFJ)",    font=F_TH, fillc=fill(NAVY), align=A_CTR)
cell(ws, f"D{row}", "Marginal rate",         font=F_TH, fillc=fill(NAVY), align=A_CTR)
brackets_start_row = row + 1

brackets = [
    (11925,  23850,  0.10),
    (48475,  96950,  0.12),
    (103350, 206700, 0.22),
    (197300, 394600, 0.24),
    (250525, 501050, 0.32),
    (626350, 751600, 0.35),
    (999999999, 999999999, 0.37),
]
for top_s, top_m, rate in brackets:
    row += 1
    cell(ws, f"B{row}", top_s, font=F_BODY, fillc=fill(GOLD_PALE), align=A_RIGHT, border=box(GOLD), fmt='"$"#,##0')
    cell(ws, f"C{row}", top_m, font=F_BODY, fillc=fill(GOLD_PALE), align=A_RIGHT, border=box(GOLD), fmt='"$"#,##0')
    cell(ws, f"D{row}", rate,  font=F_BODY, fillc=fill(GOLD_PALE), align=A_RIGHT, border=box(GOLD), fmt='0.0%')
    ws.row_dimensions[row].height = 20

brackets_end_row = row
wb.defined_names["BracketsSingle"] = DefinedName("BracketsSingle", attr_text=f"Setup!$B${brackets_start_row}:$B${brackets_end_row}")
wb.defined_names["BracketsMFJ"]    = DefinedName("BracketsMFJ",    attr_text=f"Setup!$C${brackets_start_row}:$C${brackets_end_row}")
wb.defined_names["BracketsRate"]   = DefinedName("BracketsRate",   attr_text=f"Setup!$D${brackets_start_row}:$D${brackets_end_row}")

row += 2
ws.row_dimensions[row].height = 36
cell(ws, f"B{row}",
     "HoH and MFS brackets aren't modeled separately — if you use those statuses, set filing status above to the closest match (HOH → MFJ for brackets, MFS → Single) and adjust the standard deduction. For complex returns, consult a CPA.",
     font=F_MUTED, align=A_LEFT_T)
ws.merge_cells(f"B{row}:D{row}")

# ════════════════════════════════════════════════════════════════════
# SHEET 3 — INCOME FORECAST
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Income Forecast")
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 16
for col in "CDEFGHIJ":
    ws.column_dimensions[col].width = 14

ws.row_dimensions[2].height = 32
cell(ws, "B2", "Income Forecast", font=Font(name="Calibri", size=22, bold=True, color=NAVY), align=A_LEFT)
ws.merge_cells("B2:J2")
cell(ws, "B3",
     "Net self-employment income per month (gross minus deductions like mileage). Add platforms in row 5 — up to 6 streams.",
     font=F_MUTED, align=A_LEFT)
ws.merge_cells("B3:J3")
cell(ws, "B4",
     "Hint: if you'd rather just enter one annual number, fill in \"Estimated net SE income (annual)\" on the Setup tab — that will override this forecast.",
     font=F_MUTED, align=A_LEFT)
ws.merge_cells("B4:J4")

# Header — months + platform names
ws.row_dimensions[5].height = 28
cell(ws, "B5", "Month",       font=F_TH, fillc=fill(NAVY), align=A_CTR)
cell(ws, "C5", "DoorDash",    font=F_TH, fillc=fill(NAVY), align=A_CTR)
cell(ws, "D5", "Uber / Lyft", font=F_TH, fillc=fill(NAVY), align=A_CTR)
cell(ws, "E5", "Airbnb / STR",font=F_TH, fillc=fill(NAVY), align=A_CTR)
cell(ws, "F5", "Etsy / sales",font=F_TH, fillc=fill(NAVY), align=A_CTR)
cell(ws, "G5", "Freelance",   font=F_TH, fillc=fill(NAVY), align=A_CTR)
cell(ws, "H5", "Other",       font=F_TH, fillc=fill(NAVY), align=A_CTR)
cell(ws, "I5", "Total",       font=F_TH, fillc=fill(GOLD), align=A_CTR)

months = ["January","February","March","April","May","June","July","August","September","October","November","December"]
for i, m in enumerate(months):
    r = 6 + i
    ws.row_dimensions[r].height = 22
    cell(ws, f"B{r}", m, font=F_BODY, align=A_LEFT, border=box())
    for col in "CDEFGH":
        c = ws[f"{col}{r}"]
        c.value = 0
        c.fill = fill(GOLD_PALE)
        c.font = F_BODY
        c.alignment = A_RIGHT
        c.border = box(GOLD)
        c.number_format = '"$"#,##0'
    # Row total
    c = ws[f"I{r}"]
    c.value = f"=SUM(C{r}:H{r})"
    c.font = F_CALC
    c.fill = fill(CREAM_D)
    c.alignment = A_RIGHT
    c.border = box()
    c.number_format = '"$"#,##0'

# Quarterly + annual totals row
ws.row_dimensions[18].height = 30
cell(ws, "B18", "Annual total", font=F_TOTAL, fillc=fill(GOLD_PALE), align=A_LEFT, border=box(GOLD))
for col in "CDEFGHI":
    c = ws[f"{col}18"]
    c.value = f"=SUM({col}6:{col}17)"
    c.font = F_TOTAL
    c.fill = fill(GOLD_PALE)
    c.alignment = A_RIGHT
    c.border = box(GOLD)
    c.number_format = '"$"#,##0'

# Quarterly breakdown
ws.row_dimensions[20].height = 24
cell(ws, "B20", "By quarter", font=F_H1, align=A_LEFT)
ws.merge_cells("B20:I20")

ws.row_dimensions[21].height = 24
cell(ws, "B21", "Quarter",    font=F_TH, fillc=fill(NAVY), align=A_CTR)
cell(ws, "C21", "Months",     font=F_TH, fillc=fill(NAVY), align=A_CTR)
cell(ws, "D21", "Net SE income", font=F_TH, fillc=fill(NAVY), align=A_CTR)

q_rows = [("Q1","Jan–Mar","SUM(I6:I8)"), ("Q2","Apr–May","SUM(I9:I10)"),
          ("Q3","Jun–Aug","SUM(I11:I13)"), ("Q4","Sep–Dec","SUM(I14:I17)")]
for i, (q, m, formula) in enumerate(q_rows):
    r = 22 + i
    ws.row_dimensions[r].height = 22
    cell(ws, f"B{r}", q, font=F_BODY, align=A_CTR, border=box())
    cell(ws, f"C{r}", m, font=F_BODY, align=A_LEFT, border=box())
    c = ws[f"D{r}"]
    c.value = f"={formula}"
    c.font = F_CALC
    c.alignment = A_RIGHT
    c.border = box()
    c.number_format = '"$"#,##0'

# Named ranges
wb.defined_names["Q1Income"] = DefinedName("Q1Income", attr_text="'Income Forecast'!$D$22")
wb.defined_names["Q2Income"] = DefinedName("Q2Income", attr_text="'Income Forecast'!$D$23")
wb.defined_names["Q3Income"] = DefinedName("Q3Income", attr_text="'Income Forecast'!$D$24")
wb.defined_names["Q4Income"] = DefinedName("Q4Income", attr_text="'Income Forecast'!$D$25")
wb.defined_names["AnnualSEForecast"] = DefinedName("AnnualSEForecast", attr_text="'Income Forecast'!$I$18")

# ════════════════════════════════════════════════════════════════════
# SHEET 4 — TAX CALCULATOR
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Tax Calculator")
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 44
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 60

ws.row_dimensions[2].height = 32
cell(ws, "B2", "Tax Calculator", font=Font(name="Calibri", size=22, bold=True, color=NAVY), align=A_LEFT)
ws.merge_cells("B2:D2")
cell(ws, "B3", "All numbers below are calculated — don't edit. To change inputs, use the Setup tab.",
     font=F_MUTED, align=A_LEFT)
ws.merge_cells("B3:D3")

def calc_row(r, label, formula, hint, fmt='"$"#,##0', highlight=False):
    cell(ws, f"B{r}", label, font=F_BODY if not highlight else F_TOTAL, align=A_LEFT)
    c = ws[f"C{r}"]
    c.value = formula
    c.font = F_CALC if not highlight else F_TOTAL
    c.alignment = A_RIGHT
    c.border = box()
    c.fill = fill(GOLD_PALE if highlight else CREAM_D)
    c.number_format = fmt
    cell(ws, f"D{r}", hint, font=F_MUTED, align=A_LEFT)
    ws.row_dimensions[r].height = 26 if highlight else 22

# Section: Self-employment income
ws.row_dimensions[5].height = 24
cell(ws, "B5", "Self-employment income & SE tax", font=F_H1, align=A_LEFT)
ws.merge_cells("B5:D5")

# Net SE income — use override if > 0, else forecast
calc_row(6, "Net SE income (this year)",
         "=IF(SEOverride>0,SEOverride,AnnualSEForecast)",
         "From Setup override if filled; otherwise the Income Forecast annual total")
NET_SE_CELL = "'Tax Calculator'!$C$6"

calc_row(7, "× 92.35% (SE tax base)",
         f"={NET_SE_CELL}*0.9235",
         "The IRS taxes 92.35% of SE income for Social Security & Medicare")
SE_BASE_CELL = "'Tax Calculator'!$C$7"

calc_row(8, "Social Security portion of SE tax (12.4% up to wage base)",
         f"=MIN({SE_BASE_CELL},SSWageBase)*0.124",
         "Stops at the SS wage base (set on Setup tab)")
SS_TAX_CELL = "'Tax Calculator'!$C$8"

calc_row(9, "Medicare portion of SE tax (2.9%)",
         f"={SE_BASE_CELL}*0.029",
         "No income cap on Medicare")
MEDICARE_CELL = "'Tax Calculator'!$C$9"

calc_row(10, "Total SE tax",
         f"={SS_TAX_CELL}+{MEDICARE_CELL}",
         "Sum of Social Security + Medicare", highlight=True)
SE_TAX_CELL = "'Tax Calculator'!$C$10"

calc_row(11, "Deductible half of SE tax",
         f"={SE_TAX_CELL}/2",
         "Half of SE tax reduces your AGI (it's an above-the-line deduction)")
HALF_SE_CELL = "'Tax Calculator'!$C$11"

# Section: Income tax
ws.row_dimensions[13].height = 24
cell(ws, "B13", "Federal income tax", font=F_H1, align=A_LEFT)
ws.merge_cells("B13:D13")

calc_row(14, "AGI (W-2 + net SE − half SE tax)",
         f"=W2Income+{NET_SE_CELL}-{HALF_SE_CELL}",
         "Adjusted Gross Income — the number tax brackets apply to (after standard deduction)")
AGI_CELL = "'Tax Calculator'!$C$14"

calc_row(15, "Standard deduction (based on filing status)",
         "=IF(FilingStatus=\"MFJ\",StdDedMFJ,IF(FilingStatus=\"HOH\",StdDedHOH,StdDedSingle))",
         "Auto-picks the right standard deduction; MFS uses Single")
STD_CELL = "'Tax Calculator'!$C$15"

calc_row(16, "Taxable income",
         f"=MAX({AGI_CELL}-{STD_CELL},0)",
         "AGI minus standard deduction (we don't model QBI or itemized — see CPA if relevant)")
TAXABLE_CELL = "'Tax Calculator'!$C$16"

# Federal income tax — bracket-by-bracket calculation
# Uses SUMPRODUCT trick: for each bracket, tax = max(0, min(income, top) - prev_top) * rate
# We model this with a helper approach inline.
# Single brackets: 11925/48475/103350/197300/250525/626350/999999999 with rates 10/12/22/24/32/35/37
# MFJ: 23850/96950/206700/394600/501050/751600/999999999

fed_tax_formula = (
    '=IF(FilingStatus="MFJ",'
    'SUMPRODUCT(((MIN(' + TAXABLE_CELL + ',BracketsMFJ)-IFERROR(OFFSET(BracketsMFJ,-1,0),0))>0)'
    '*(MIN(' + TAXABLE_CELL + ',BracketsMFJ)-IFERROR(OFFSET(BracketsMFJ,-1,0),0))*BracketsRate),'
    'SUMPRODUCT(((MIN(' + TAXABLE_CELL + ',BracketsSingle)-IFERROR(OFFSET(BracketsSingle,-1,0),0))>0)'
    '*(MIN(' + TAXABLE_CELL + ',BracketsSingle)-IFERROR(OFFSET(BracketsSingle,-1,0),0))*BracketsRate)'
    ')'
)
calc_row(17, "Federal income tax (bracket calc)",
         fed_tax_formula,
         "Applied bracket-by-bracket to your taxable income", highlight=True)
FED_TAX_CELL = "'Tax Calculator'!$C$17"

# Section: State + total
ws.row_dimensions[19].height = 24
cell(ws, "B19", "State tax & totals", font=F_H1, align=A_LEFT)
ws.merge_cells("B19:D19")

calc_row(20, "State income tax (flat rate × taxable income)",
         f"={TAXABLE_CELL}*StateRate",
         "Rough estimate — actual state tax varies by bracket. Override the rate on Setup if you have a better number.", highlight=True)
STATE_TAX_CELL = "'Tax Calculator'!$C$20"

calc_row(21, "Total federal owed (SE + income)",
         f"={SE_TAX_CELL}+{FED_TAX_CELL}",
         "What you owe the IRS this year, before W-2 withholding")
TOTAL_FED_CELL = "'Tax Calculator'!$C$21"

calc_row(22, "Less: W-2 federal withholding",
         "=-W2Withheld",
         "Subtracts what your employer already sent in for you")

calc_row(23, "= Net federal owed (quarterly target)",
         f"={TOTAL_FED_CELL}-W2Withheld",
         "Divide this by 4 to get your quarterly federal payment", highlight=True)
NET_FED_CELL = "'Tax Calculator'!$C$23"

calc_row(24, "Total state owed (quarterly target)",
         f"={STATE_TAX_CELL}",
         "Divide this by 4 to get your quarterly state payment", highlight=True)
NET_STATE_CELL = "'Tax Calculator'!$C$24"

# Named ranges
wb.defined_names["NetFedOwed"]   = DefinedName("NetFedOwed",   attr_text=NET_FED_CELL)
wb.defined_names["NetStateOwed"] = DefinedName("NetStateOwed", attr_text=NET_STATE_CELL)
wb.defined_names["TotalFedTax"]  = DefinedName("TotalFedTax",  attr_text=TOTAL_FED_CELL)
wb.defined_names["SETax"]        = DefinedName("SETax",        attr_text=SE_TAX_CELL)
wb.defined_names["FedIncomeTax"] = DefinedName("FedIncomeTax", attr_text=FED_TAX_CELL)

# ════════════════════════════════════════════════════════════════════
# SHEET 5 — QUARTERLY PAYMENTS
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Quarterly Payments")
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 18
ws.column_dimensions["G"].width = 18
ws.column_dimensions["H"].width = 20

ws.row_dimensions[2].height = 32
cell(ws, "B2", "Quarterly Payments", font=Font(name="Calibri", size=22, bold=True, color=NAVY), align=A_LEFT)
ws.merge_cells("B2:H2")
cell(ws, "B3", "Log each payment after you make it. Status colors track whether you're on pace to hit your annual target.",
     font=F_MUTED, align=A_LEFT)
ws.merge_cells("B3:H3")

# Federal table
ws.row_dimensions[5].height = 24
cell(ws, "B5", "Federal (IRS)", font=F_H1, align=A_LEFT)
ws.merge_cells("B5:H5")

headers = ["Q", "Due date", "Target $", "Paid date", "Paid $", "Method", "Status"]
ws.row_dimensions[6].height = 24
for i, h in enumerate(headers):
    c = ws.cell(row=6, column=2+i, value=h)
    c.font = F_TH
    c.fill = fill(NAVY)
    c.alignment = A_CTR
    c.border = box(NAVY)

# Q1-Q4 federal — target = NetFedOwed / 4
due_dates = [("Q1","April 15"), ("Q2","June 15"), ("Q3","September 15"), ("Q4","January 15")]
for i, (q, dd) in enumerate(due_dates):
    r = 7 + i
    ws.row_dimensions[r].height = 24
    cell(ws, f"B{r}", q,  font=F_BODY, align=A_CTR, border=box())
    cell(ws, f"C{r}", dd, font=F_BODY, align=A_LEFT, border=box())
    # Target $
    c = ws[f"D{r}"]
    c.value = "=NetFedOwed/4"
    c.font = F_CALC
    c.alignment = A_RIGHT
    c.border = box()
    c.number_format = '"$"#,##0'
    c.fill = fill(CREAM_D)
    # Paid date / paid $ / method — user inputs
    for col in "EFG":
        cc = ws[f"{col}{r}"]
        cc.fill = fill(GOLD_PALE)
        cc.border = box(GOLD)
        cc.font = F_BODY
        if col == "E":
            cc.number_format = "yyyy-mm-dd"
        elif col == "F":
            cc.number_format = '"$"#,##0'
            cc.alignment = A_RIGHT
        else:
            cc.alignment = A_LEFT
    # Status formula
    status = ws[f"H{r}"]
    status.value = (
        f'=IF(F{r}="","Not paid yet",'
        f'IF(F{r}>=D{r}*0.95,"On track",'
        f'"Underpaid by $"&TEXT(D{r}-F{r},"#,##0")))'
    )
    status.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    status.alignment = A_LEFT
    status.border = box()

# Totals row federal
ws.row_dimensions[11].height = 28
cell(ws, "B11", "Total", font=F_TOTAL, fillc=fill(GOLD_PALE), align=A_LEFT, border=box(GOLD))
ws.merge_cells("B11:C11")
ws["D11"].value = "=SUM(D7:D10)"
ws["D11"].font = F_TOTAL
ws["D11"].fill = fill(GOLD_PALE)
ws["D11"].alignment = A_RIGHT
ws["D11"].border = box(GOLD)
ws["D11"].number_format = '"$"#,##0'
ws["F11"].value = "=SUM(F7:F10)"
ws["F11"].font = F_TOTAL
ws["F11"].fill = fill(GOLD_PALE)
ws["F11"].alignment = A_RIGHT
ws["F11"].border = box(GOLD)
ws["F11"].number_format = '"$"#,##0'
ws["H11"].value = '=IF(F11>=D11*0.95,"Annual target met","Short $"&TEXT(D11-F11,"#,##0"))'
ws["H11"].font = F_TOTAL
ws["H11"].fill = fill(GOLD_PALE)
ws["H11"].alignment = A_LEFT
ws["H11"].border = box(GOLD)

# Status color rules
for r in range(7, 11):
    ws.conditional_formatting.add(f"H{r}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("On track",H{r}))'], fill=fill(GREEN_BG), font=Font(color=GREEN, bold=True)))
    ws.conditional_formatting.add(f"H{r}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("Not paid",H{r}))'], fill=fill(AMBER_BG), font=Font(color=AMBER, bold=True)))
    ws.conditional_formatting.add(f"H{r}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("Underpaid",H{r}))'], fill=fill(RED_BG), font=Font(color=RED, bold=True)))
ws.conditional_formatting.add("H11",
    FormulaRule(formula=['ISNUMBER(SEARCH("met",H11))'], fill=fill(GREEN_BG), font=Font(color=GREEN, bold=True)))
ws.conditional_formatting.add("H11",
    FormulaRule(formula=['ISNUMBER(SEARCH("Short",H11))'], fill=fill(RED_BG), font=Font(color=RED, bold=True)))

# State table
ws.row_dimensions[14].height = 24
cell(ws, "B14", "State", font=F_H1, align=A_LEFT)
ws.merge_cells("B14:H14")

ws.row_dimensions[15].height = 24
for i, h in enumerate(headers):
    c = ws.cell(row=15, column=2+i, value=h)
    c.font = F_TH
    c.fill = fill(NAVY)
    c.alignment = A_CTR
    c.border = box(NAVY)

for i, (q, dd) in enumerate(due_dates):
    r = 16 + i
    ws.row_dimensions[r].height = 24
    cell(ws, f"B{r}", q,  font=F_BODY, align=A_CTR, border=box())
    cell(ws, f"C{r}", dd, font=F_BODY, align=A_LEFT, border=box())
    c = ws[f"D{r}"]
    c.value = "=NetStateOwed/4"
    c.font = F_CALC
    c.alignment = A_RIGHT
    c.border = box()
    c.number_format = '"$"#,##0'
    c.fill = fill(CREAM_D)
    for col in "EFG":
        cc = ws[f"{col}{r}"]
        cc.fill = fill(GOLD_PALE)
        cc.border = box(GOLD)
        cc.font = F_BODY
        if col == "E":   cc.number_format = "yyyy-mm-dd"
        elif col == "F": cc.number_format = '"$"#,##0'; cc.alignment = A_RIGHT
        else:            cc.alignment = A_LEFT
    status = ws[f"H{r}"]
    status.value = (
        f'=IF(F{r}="","Not paid yet",'
        f'IF(F{r}>=D{r}*0.95,"On track",'
        f'"Underpaid by $"&TEXT(D{r}-F{r},"#,##0")))'
    )
    status.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    status.alignment = A_LEFT
    status.border = box()
    # Conditional formatting reuse
    ws.conditional_formatting.add(f"H{r}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("On track",H{r}))'], fill=fill(GREEN_BG), font=Font(color=GREEN, bold=True)))
    ws.conditional_formatting.add(f"H{r}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("Not paid",H{r}))'], fill=fill(AMBER_BG), font=Font(color=AMBER, bold=True)))
    ws.conditional_formatting.add(f"H{r}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("Underpaid",H{r}))'], fill=fill(RED_BG), font=Font(color=RED, bold=True)))

# State totals
ws.row_dimensions[20].height = 28
cell(ws, "B20", "Total", font=F_TOTAL, fillc=fill(GOLD_PALE), align=A_LEFT, border=box(GOLD))
ws.merge_cells("B20:C20")
ws["D20"].value = "=SUM(D16:D19)"
ws["D20"].font = F_TOTAL; ws["D20"].fill = fill(GOLD_PALE); ws["D20"].alignment = A_RIGHT; ws["D20"].border = box(GOLD); ws["D20"].number_format = '"$"#,##0'
ws["F20"].value = "=SUM(F16:F19)"
ws["F20"].font = F_TOTAL; ws["F20"].fill = fill(GOLD_PALE); ws["F20"].alignment = A_RIGHT; ws["F20"].border = box(GOLD); ws["F20"].number_format = '"$"#,##0'

# Named ranges for safe harbor sheet
wb.defined_names["FedPaidYTD"]   = DefinedName("FedPaidYTD",   attr_text="'Quarterly Payments'!$F$11")
wb.defined_names["StatePaidYTD"] = DefinedName("StatePaidYTD", attr_text="'Quarterly Payments'!$F$20")

# ════════════════════════════════════════════════════════════════════
# SHEET 6 — SAFE HARBOR CHECK
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Safe Harbor")
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 45

ws.row_dimensions[2].height = 32
cell(ws, "B2", "Safe Harbor Check", font=Font(name="Calibri", size=22, bold=True, color=NAVY), align=A_LEFT)
ws.merge_cells("B2:D2")
cell(ws, "B3",
     "The IRS won't charge you an underpayment penalty if you hit EITHER of these two safe-harbor tests. You only need ONE.",
     font=F_MUTED, align=A_LEFT)
ws.merge_cells("B3:D3")

# Test 1: 100% / 110% of last year
ws.row_dimensions[5].height = 24
cell(ws, "B5", "Safe Harbor 1 — based on last year's tax", font=F_H1, align=A_LEFT)
ws.merge_cells("B5:D5")

calc_row_sh = lambda r, label, formula, hint, fmt='"$"#,##0', hl=False: (
    cell(ws, f"B{r}", label, font=F_BODY if not hl else F_TOTAL, align=A_LEFT),
    setattr(ws[f"C{r}"], "value", formula),
    setattr(ws[f"C{r}"], "font", F_CALC if not hl else F_TOTAL),
    setattr(ws[f"C{r}"], "alignment", A_RIGHT),
    setattr(ws[f"C{r}"], "border", box()),
    setattr(ws[f"C{r}"], "number_format", fmt),
    setattr(ws[f"C{r}"], "fill", fill(GOLD_PALE if hl else CREAM_D)),
    cell(ws, f"D{r}", hint, font=F_MUTED, align=A_LEFT),
)

cell(ws, "B6", "Required multiplier (110% if last year AGI > $150K, else 100%)", font=F_BODY, align=A_LEFT)
c = ws["C6"]; c.value = "=IF(LastYearAGI>150000,1.1,1)"; c.font = F_CALC; c.alignment = A_RIGHT; c.border = box(); c.number_format = "0%"; c.fill = fill(CREAM_D)
cell(ws, "D6", "The 110% rule kicks in for higher earners", font=F_MUTED, align=A_LEFT)
ws.row_dimensions[6].height = 22

cell(ws, "B7", "Safe harbor target (last year's tax × multiplier)", font=F_TOTAL, align=A_LEFT)
c = ws["C7"]; c.value = "=LastYearTax*C6"; c.font = F_TOTAL; c.alignment = A_RIGHT; c.border = box(GOLD); c.number_format = '"$"#,##0'; c.fill = fill(GOLD_PALE)
cell(ws, "D7", "If you pay AT LEAST this much across all 4 quarters + W-2 withholding, you're safe", font=F_MUTED, align=A_LEFT)
ws.row_dimensions[7].height = 26

cell(ws, "B8", "Paid so far (Q1–Q4 federal + W-2 withholding)", font=F_BODY, align=A_LEFT)
c = ws["C8"]; c.value = "=FedPaidYTD+W2Withheld"; c.font = F_CALC; c.alignment = A_RIGHT; c.border = box(); c.number_format = '"$"#,##0'; c.fill = fill(CREAM_D)
cell(ws, "D8", "From Quarterly Payments tab + W-2 withholding on Setup", font=F_MUTED, align=A_LEFT)
ws.row_dimensions[8].height = 22

cell(ws, "B9", "Safe Harbor 1 status:", font=F_TOTAL, align=A_LEFT)
c = ws["C9"]
c.value = '=IF(C8>=C7,"✓ SAFE","Short $"&TEXT(C7-C8,"#,##0"))'
c.font = Font(name="Calibri", size=14, bold=True, color=NAVY)
c.alignment = A_CTR
c.border = box(GOLD)
ws.row_dimensions[9].height = 32
ws.conditional_formatting.add("C9",
    FormulaRule(formula=['ISNUMBER(SEARCH("SAFE",C9))'], fill=fill(GREEN_BG), font=Font(color=GREEN, bold=True, size=14)))
ws.conditional_formatting.add("C9",
    FormulaRule(formula=['ISNUMBER(SEARCH("Short",C9))'], fill=fill(RED_BG), font=Font(color=RED, bold=True, size=14)))

# Test 2: 90% of current year
ws.row_dimensions[11].height = 24
cell(ws, "B11", "Safe Harbor 2 — based on this year's tax", font=F_H1, align=A_LEFT)
ws.merge_cells("B11:D11")

cell(ws, "B12", "This year's projected total federal tax", font=F_BODY, align=A_LEFT)
c = ws["C12"]; c.value = "=TotalFedTax"; c.font = F_CALC; c.alignment = A_RIGHT; c.border = box(); c.number_format = '"$"#,##0'; c.fill = fill(CREAM_D)
cell(ws, "D12", "From Tax Calculator", font=F_MUTED, align=A_LEFT)
ws.row_dimensions[12].height = 22

cell(ws, "B13", "Safe harbor target (90% of this year's tax)", font=F_TOTAL, align=A_LEFT)
c = ws["C13"]; c.value = "=C12*0.9"; c.font = F_TOTAL; c.alignment = A_RIGHT; c.border = box(GOLD); c.number_format = '"$"#,##0'; c.fill = fill(GOLD_PALE)
cell(ws, "D13", "If you pay at least 90% of what you'll actually owe, you're safe", font=F_MUTED, align=A_LEFT)
ws.row_dimensions[13].height = 26

cell(ws, "B14", "Paid so far (Q1–Q4 federal + W-2 withholding)", font=F_BODY, align=A_LEFT)
c = ws["C14"]; c.value = "=FedPaidYTD+W2Withheld"; c.font = F_CALC; c.alignment = A_RIGHT; c.border = box(); c.number_format = '"$"#,##0'; c.fill = fill(CREAM_D)
cell(ws, "D14", "Same total as Safe Harbor 1", font=F_MUTED, align=A_LEFT)
ws.row_dimensions[14].height = 22

cell(ws, "B15", "Safe Harbor 2 status:", font=F_TOTAL, align=A_LEFT)
c = ws["C15"]
c.value = '=IF(C14>=C13,"✓ SAFE","Short $"&TEXT(C13-C14,"#,##0"))'
c.font = Font(name="Calibri", size=14, bold=True, color=NAVY)
c.alignment = A_CTR
c.border = box(GOLD)
ws.row_dimensions[15].height = 32
ws.conditional_formatting.add("C15",
    FormulaRule(formula=['ISNUMBER(SEARCH("SAFE",C15))'], fill=fill(GREEN_BG), font=Font(color=GREEN, bold=True, size=14)))
ws.conditional_formatting.add("C15",
    FormulaRule(formula=['ISNUMBER(SEARCH("Short",C15))'], fill=fill(RED_BG), font=Font(color=RED, bold=True, size=14)))

# Bottom line
ws.row_dimensions[17].height = 24
cell(ws, "B17", "Are you penalty-safe?", font=F_H1, align=A_LEFT)
ws.merge_cells("B17:D17")

cell(ws, "B18", "Overall status:", font=F_TOTAL, align=A_LEFT)
c = ws["C18"]
c.value = ('=IF(OR(C8>=C7,C14>=C13),"✓ YES — you\'re safe from penalty",'
           '"⚠ Not yet — close one of the gaps above")')
c.font = Font(name="Calibri", size=14, bold=True, color=NAVY)
c.alignment = A_CTR
c.border = box(GOLD)
ws.row_dimensions[18].height = 38
ws.conditional_formatting.add("C18",
    FormulaRule(formula=['ISNUMBER(SEARCH("YES",C18))'], fill=fill(GREEN_BG), font=Font(color=GREEN, bold=True, size=14)))
ws.conditional_formatting.add("C18",
    FormulaRule(formula=['ISNUMBER(SEARCH("Not yet",C18))'], fill=fill(AMBER_BG), font=Font(color=AMBER, bold=True, size=14)))

# Plain-English note
ws.row_dimensions[20].height = 70
note = ("You only need to satisfy ONE of the two tests to avoid the underpayment penalty. Most "
        "side hustlers hit Safe Harbor 1 because it's easier — you just need to pay last year's tax bill "
        "across four equal quarters. Safe Harbor 2 is for years when your income drops significantly.")
cell(ws, "B20", note, font=F_BODY, align=A_LEFT_T)
ws.merge_cells("B20:D20")

# ════════════════════════════════════════════════════════════════════
# SHEET 7 — IRS & STATE REFERENCE
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("IRS Reference")
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 70

ws.row_dimensions[2].height = 32
cell(ws, "B2", "IRS & State Quarterly Tax Reference", font=Font(name="Calibri", size=22, bold=True, color=NAVY), align=A_LEFT)
ws.merge_cells("B2:C2")
cell(ws, "B3", "Plain-English answers to what gets people in trouble.", font=F_MUTED, align=A_LEFT)
ws.merge_cells("B3:C3")

faqs = [
    ("Where do I actually pay federal quarterly tax?",
     "Three options. (1) IRS Direct Pay — irs.gov/payments/direct-pay — free, takes 2 minutes, pulls from a bank account. (2) EFTPS — eftps.gov — also free, requires a one-time enrollment, better if you want a permanent payment history. (3) Mail a check with Form 1040-ES voucher to the IRS address listed on the form for your state. Direct Pay is the easiest for almost everyone."),
    ("Do I need to file Form 1040-ES?",
     "Technically no — Form 1040-ES is the paper voucher you mail with a check. If you pay electronically via Direct Pay or EFTPS, you don't file anything; the payment IS the filing. Just keep your confirmation number with your tax records."),
    ("How does the state work?",
     "Every state with income tax has its own quarterly system. Most (CA, NY, IL, etc.) mirror the federal due dates. Search \"[your state] estimated tax payments\" — usually it's a one-page payment portal on your state's revenue/tax department site. California is FTB Web Pay. New York is the IT-2105 voucher or online via NY.gov. Texas, Florida, Washington, Nevada, Tennessee, South Dakota, Wyoming, Alaska, and New Hampshire have no state income tax — skip this entirely."),
    ("What if my income is uneven across the year?",
     "You can use the \"annualized income installment method\" (Form 2210, Schedule AI) to pay less in slow quarters and more in busy ones — but it's a pain. Most people just pay equal quarters using Safe Harbor 1 (100% / 110% of last year's tax), which doesn't care when the income actually comes in."),
    ("What's the actual penalty if I skip a quarter?",
     "The IRS charges the federal short-term rate + 3% on the underpaid amount, calculated quarterly. In 2025 that worked out to ~8% annualized. So if you should have paid $3,000 for Q1 and didn't, by April you owe an extra ~$60. Skip all four quarters on a $15K bill and you're looking at $400-$800 in penalties. The penalty is non-deductible — pure waste."),
    ("Can I overpay to be safe?",
     "Yes, and many people do. Pay an extra 5-10% per quarter to give yourself buffer; the IRS refunds the excess (or rolls it forward) when you file. There's no penalty for overpaying. The only \"cost\" is the lost interest on that cash — usually pennies compared to the underpayment penalty."),
    ("I had a huge year. Last year was tiny. Do I really only owe what last year's tax was × 4?",
     "Under Safe Harbor 1, yes — for THIS year. You hit the safe harbor by paying 100%/110% of last year's tax. The catch: in April when you file, you'll owe a LARGE balance for this year's true tax. No penalty, but the bill is still due. Start putting cash aside in a separate account so April doesn't bury you."),
    ("What if I miss a quarter and want to catch up?",
     "Pay the missed amount as soon as you can with your next payment. The penalty accrues per day, so paying late is much better than not paying. You can't \"backdate\" a quarter, but you can stop the bleeding."),
    ("Does this work if I'm both W-2 and self-employed?",
     "Yes. The Tax Calculator already factors in your W-2 income and withholding. The Net federal owed number on that tab assumes your employer's withholding takes care of the W-2 portion, so you only need to send quarterlies for the SE side. Make sure the W-2 withholding number you entered is the FULL YEAR amount, not just YTD."),
    ("Do I need to keep records of payments?",
     "Yes — keep electronic confirmations (Direct Pay sends an email; EFTPS keeps history online; mailed checks → keep the bank's cleared-check image). Match these to the Paid date and Paid $ columns on the Quarterly Payments tab. At year-end, the total of your four quarterly payments goes on Form 1040 line 26."),
]

row = 5
for q, a in faqs:
    ws.row_dimensions[row].height = 22
    cell(ws, f"B{row}", "Q", font=Font(name="Calibri", size=12, bold=True, color=GOLD), align=A_CTR)
    cell(ws, f"C{row}", q, font=F_H2, align=A_LEFT)
    row += 1
    ws.row_dimensions[row].height = 70
    cell(ws, f"C{row}", a, font=F_BODY, align=A_LEFT_T)
    row += 2  # spacer

# ════════════════════════════════════════════════════════════════════
# SHEET 8 — DASHBOARD
# (created LAST in code but moved to position 2 via _sheets reorder below)
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Dashboard")
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110

ws.column_dimensions["A"].width = 2
for col, w in zip("BCDEFGHIJ", [20, 20, 20, 20, 2, 20, 20, 20, 20]):
    ws.column_dimensions[col].width = w

ws.row_dimensions[2].height = 36
cell(ws, "B2", "Dashboard", font=Font(name="Calibri", size=24, bold=True, color=NAVY), align=A_LEFT)
ws.merge_cells("B2:J2")
cell(ws, "B3", "One screen — where you stand right now. All values update as you fill in the other tabs.",
     font=F_MUTED, align=A_LEFT)
ws.merge_cells("B3:J3")

# ── KPI ROW 1 ──
def kpi_card(row, col_start, label, formula, fmt, value_color=NAVY, fill_color=CREAM):
    end_col_letter = get_column_letter(openpyxl.utils.column_index_from_string(col_start) + 1)
    # Label cell
    lbl = ws[f"{col_start}{row}"]
    lbl.value = label
    lbl.font = F_KPI_LBL
    lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lbl.fill = fill(fill_color)
    lbl.border = Border(
        left=Side(style="thin", color=BORDER),
        right=Side(style="thin", color=BORDER),
        top=Side(style="thin", color=BORDER),
        bottom=Side(style=None)
    )
    # Make label cell span 2 cols
    ws.merge_cells(start_row=row, start_column=openpyxl.utils.column_index_from_string(col_start),
                   end_row=row, end_column=openpyxl.utils.column_index_from_string(col_start)+1)
    # Value cell
    val = ws[f"{col_start}{row+1}"]
    val.value = formula
    val.font = Font(name="Calibri", size=22, bold=True, color=value_color)
    val.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    val.fill = fill(fill_color)
    val.number_format = fmt
    val.border = Border(
        left=Side(style="thin", color=BORDER),
        right=Side(style="thin", color=BORDER),
        top=Side(style=None),
        bottom=Side(style="thin", color=BORDER)
    )
    ws.merge_cells(start_row=row+1, start_column=openpyxl.utils.column_index_from_string(col_start),
                   end_row=row+1, end_column=openpyxl.utils.column_index_from_string(col_start)+1)
    ws.row_dimensions[row].height = 24
    ws.row_dimensions[row+1].height = 40

# Row 1: Annual totals
ws.row_dimensions[5].height = 24
cell(ws, "B5", "ANNUAL TAX PICTURE", font=Font(name="Calibri", size=11, bold=True, color=GOLD), align=A_LEFT)
ws.merge_cells("B5:J5")

kpi_card(6, "B", "TOTAL SE TAX",          "=SETax",        '"$"#,##0', value_color=NAVY)
kpi_card(6, "D", "FEDERAL INCOME TAX",    "=FedIncomeTax", '"$"#,##0', value_color=NAVY)
kpi_card(6, "G", "STATE TAX",             "='Tax Calculator'!$C$20", '"$"#,##0', value_color=NAVY)
kpi_card(6, "I", "TOTAL ANNUAL OWED",     "=SETax+FedIncomeTax+'Tax Calculator'!$C$20", '"$"#,##0', value_color="FFC9973A", fill_color=GOLD_PALE)

# Row 2: Payment progress
ws.row_dimensions[9].height = 12
ws.row_dimensions[10].height = 24
cell(ws, "B10", "WHERE YOU STAND TODAY", font=Font(name="Calibri", size=11, bold=True, color=GOLD), align=A_LEFT)
ws.merge_cells("B10:J10")

kpi_card(11, "B", "FED PAID YTD",         "=FedPaidYTD",   '"$"#,##0')
kpi_card(11, "D", "STATE PAID YTD",       "=StatePaidYTD", '"$"#,##0')
kpi_card(11, "G", "FEDERAL TARGET (annual)", "=NetFedOwed", '"$"#,##0')
kpi_card(11, "I", "REMAINING FEDERAL",    "=MAX(NetFedOwed-FedPaidYTD,0)", '"$"#,##0', value_color="FFB83232", fill_color=RED_BG)

# Row 3: Safe harbor + next deadline
ws.row_dimensions[14].height = 12
ws.row_dimensions[15].height = 24
cell(ws, "B15", "SAFETY & NEXT DEADLINE", font=Font(name="Calibri", size=11, bold=True, color=GOLD), align=A_LEFT)
ws.merge_cells("B15:J15")

# Safe harbor status card (spans 4 cols)
ws.row_dimensions[16].height = 24
cell(ws, "B16", "PENALTY-SAFE STATUS", font=F_KPI_LBL, fillc=fill(CREAM), align=Alignment(horizontal="left", vertical="center", indent=1),
     border=Border(left=Side(style="thin",color=BORDER), right=Side(style="thin",color=BORDER), top=Side(style="thin",color=BORDER)))
ws.merge_cells("B16:E16")
ws.row_dimensions[17].height = 44
sh_val = ws["B17"]
sh_val.value = '=IF(OR(\'Safe Harbor\'!$C$8>=\'Safe Harbor\'!$C$7,\'Safe Harbor\'!$C$14>=\'Safe Harbor\'!$C$13),"✓ Penalty-safe","⚠ Not yet penalty-safe")'
sh_val.font = Font(name="Calibri", size=16, bold=True, color=NAVY)
sh_val.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
sh_val.fill = fill(CREAM)
sh_val.border = Border(left=Side(style="thin",color=BORDER), right=Side(style="thin",color=BORDER), bottom=Side(style="thin",color=BORDER))
ws.merge_cells("B17:E17")
ws.conditional_formatting.add("B17",
    FormulaRule(formula=['ISNUMBER(SEARCH("safe",B17))'], fill=fill(GREEN_BG), font=Font(name="Calibri", size=16, bold=True, color=GREEN)))
ws.conditional_formatting.add("B17",
    FormulaRule(formula=['ISNUMBER(SEARCH("Not yet",B17))'], fill=fill(AMBER_BG), font=Font(name="Calibri", size=16, bold=True, color=AMBER)))

# Next deadline card (spans 4 cols)
cell(ws, "G16", "NEXT QUARTERLY DEADLINE", font=F_KPI_LBL, fillc=fill(CREAM), align=Alignment(horizontal="left", vertical="center", indent=1),
     border=Border(left=Side(style="thin",color=BORDER), right=Side(style="thin",color=BORDER), top=Side(style="thin",color=BORDER)))
ws.merge_cells("G16:J16")
nd_val = ws["G17"]
# Computes the next upcoming Q deadline by comparing TODAY() to the four dates
nd_val.value = (
    '=IF(TODAY()<=DATE(TaxYear,4,15),"Q1 — April 15, "&TEXT(DATE(TaxYear,4,15)-TODAY(),"0")&" days away",'
    'IF(TODAY()<=DATE(TaxYear,6,15),"Q2 — June 15, "&TEXT(DATE(TaxYear,6,15)-TODAY(),"0")&" days away",'
    'IF(TODAY()<=DATE(TaxYear,9,15),"Q3 — September 15, "&TEXT(DATE(TaxYear,9,15)-TODAY(),"0")&" days away",'
    'IF(TODAY()<=DATE(TaxYear+1,1,15),"Q4 — January 15 (next year), "&TEXT(DATE(TaxYear+1,1,15)-TODAY(),"0")&" days away",'
    '"All quarterly deadlines passed for "&TaxYear))))'
)
nd_val.font = Font(name="Calibri", size=14, bold=True, color=NAVY)
nd_val.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
nd_val.fill = fill(CREAM)
nd_val.border = Border(left=Side(style="thin",color=BORDER), right=Side(style="thin",color=BORDER), bottom=Side(style="thin",color=BORDER))
ws.merge_cells("G17:J17")

# ── CHART: Quarterly target vs paid ──
ws.row_dimensions[19].height = 12
ws.row_dimensions[20].height = 24
cell(ws, "B20", "QUARTERLY PROGRESS — FEDERAL", font=Font(name="Calibri", size=11, bold=True, color=GOLD), align=A_LEFT)
ws.merge_cells("B20:J20")

# Hidden data block for the chart (rows 22-26, cols B-D)
ws.row_dimensions[22].height = 18
cell(ws, "B22", "Quarter", font=F_BODY, align=A_LEFT)
cell(ws, "C22", "Target $", font=F_BODY, align=A_RIGHT)
cell(ws, "D22", "Paid $",   font=F_BODY, align=A_RIGHT)
for i, q in enumerate(["Q1","Q2","Q3","Q4"]):
    r = 23 + i
    ws.row_dimensions[r].height = 18
    cell(ws, f"B{r}", q, font=F_BODY, align=A_LEFT)
    c = ws[f"C{r}"]; c.value = f"='Quarterly Payments'!$D${7+i}"; c.font = F_BODY; c.alignment = A_RIGHT; c.number_format = '"$"#,##0'
    c = ws[f"D{r}"]; c.value = f"='Quarterly Payments'!$F${7+i}"; c.font = F_BODY; c.alignment = A_RIGHT; c.number_format = '"$"#,##0'

# Create chart
chart = BarChart()
chart.type = "col"
chart.style = 11
chart.title = "Federal — Target vs Paid by Quarter"
chart.y_axis.title = "Dollars"
chart.x_axis.title = "Quarter"
chart.height = 8
chart.width = 18

target_data = Reference(ws, min_col=3, min_row=22, max_col=3, max_row=26)
paid_data   = Reference(ws, min_col=4, min_row=22, max_col=4, max_row=26)
cats        = Reference(ws, min_col=2, min_row=23, max_col=2, max_row=26)
chart.add_data(target_data, titles_from_data=True)
chart.add_data(paid_data,   titles_from_data=True)
chart.set_categories(cats)
chart.dataLabels = DataLabelList(showVal=False)

# Colors — series 1 (target) navy, series 2 (paid) gold
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import ColorChoice
chart.series[0].graphicalProperties = GraphicalProperties(solidFill="1C2B4A")
chart.series[1].graphicalProperties = GraphicalProperties(solidFill="C9973A")

ws.add_chart(chart, "F22")

# Hide the data feed rows from view by setting white-on-white isn't reliable; we just keep them — they look like a tiny table.

# Footer note
ws.row_dimensions[40].height = 40
note = ("Everything on this Dashboard is calculated from the Setup, Income Forecast, and Quarterly Payments tabs. "
        "If a number looks wrong, the source is on those tabs — not here.")
cell(ws, "B40", note, font=F_MUTED, align=A_LEFT_T)
ws.merge_cells("B40:J40")

# ════════════════════════════════════════════════════════════════════
# SHEET 9 — EXAMPLES
# ════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Examples")
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110

ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 18

ws.row_dimensions[2].height = 32
cell(ws, "B2", "Worked Examples", font=Font(name="Calibri", size=22, bold=True, color=NAVY), align=A_LEFT)
ws.merge_cells("B2:E2")
cell(ws, "B3", "Three real scenarios showing exactly what numbers go where. Compare to your own situation, then plug yours into the Setup tab.",
     font=F_MUTED, align=A_LEFT)
ws.merge_cells("B3:E3")

# Headers for the 3-column comparison
ws.row_dimensions[5].height = 28
cell(ws, "B5", "", font=F_TH, fillc=fill(NAVY))
cell(ws, "C5", "DoorDash full-timer", font=F_TH, fillc=fill(NAVY), align=A_CTR)
cell(ws, "D5", "Freelancer + W-2 day job", font=F_TH, fillc=fill(NAVY), align=A_CTR)
cell(ws, "E5", "Etsy seller", font=F_TH, fillc=fill(NAVY), align=A_CTR)

# Row 6: Profile description
ws.row_dimensions[6].height = 68
cell(ws, "B6", "Situation", font=F_H2, align=A_LEFT_T)
cell(ws, "C6", "Single, lives in CA, Dashes full time. Tracked 22,000 business miles. Gross earnings $58,000 minus $15,400 mileage deduction = $42,600 net.", font=F_BODY, align=A_LEFT_T, fillc=fill(CREAM))
cell(ws, "D6", "MFJ, NY, $48K W-2 day job with $5,400 federal withholding + $60K net freelance design work on the side.", font=F_BODY, align=A_LEFT_T, fillc=fill(CREAM))
cell(ws, "E6", "Single, TX (no state tax), Etsy print shop. $25,000 net profit after COGS and home office deduction.", font=F_BODY, align=A_LEFT_T, fillc=fill(CREAM))

# Helper to add a row
def ex_row(r, label, c_val, d_val, e_val, fmt=None, height=22, is_total=False):
    cell(ws, f"B{r}", label, font=(F_TOTAL if is_total else F_BODY), align=A_LEFT, border=box(),
         fillc=fill(GOLD_PALE if is_total else None) if is_total else None)
    for col, v in zip("CDE", [c_val, d_val, e_val]):
        c = ws[f"{col}{r}"]
        c.value = v
        c.font = F_TOTAL if is_total else F_BODY
        c.alignment = A_RIGHT
        c.border = box()
        if is_total:
            c.fill = fill(GOLD_PALE)
        if fmt:
            c.number_format = fmt
    ws.row_dimensions[r].height = height

# Section A: Setup inputs
ws.row_dimensions[8].height = 24
cell(ws, "B8", "Setup tab inputs", font=F_H1, align=A_LEFT)
ws.merge_cells("B8:E8")

ex_row(9,  "Filing status",                          "Single", "MFJ", "Single")
ex_row(10, "State",                                  "CA",     "NY",  "TX")
ex_row(11, "State flat estimated tax rate",          0.06,     0.05,  0.0,    fmt="0.0%")
ex_row(12, "Net SE income (annual)",                 42600,    60000, 25000,  fmt='"$"#,##0')
ex_row(13, "W-2 income",                             0,        48000, 0,      fmt='"$"#,##0')
ex_row(14, "Federal tax withheld from W-2",          0,        5400,  0,      fmt='"$"#,##0')
ex_row(15, "Last year's total federal tax",          3200,     8900,  1800,   fmt='"$"#,##0')
ex_row(16, "Last year's AGI",                        38000,    98000, 24000,  fmt='"$"#,##0')

# Section B: What the calculator returns
ws.row_dimensions[18].height = 24
cell(ws, "B18", "What the Tax Calculator returns", font=F_H1, align=A_LEFT)
ws.merge_cells("B18:E18")

ex_row(19, "Net SE income × 92.35% (SE base)",       39341,    55410, 23088,  fmt='"$"#,##0')
ex_row(20, "Social Security (12.4%)",                4878,     6871,  2863,   fmt='"$"#,##0')
ex_row(21, "Medicare (2.9%)",                        1141,     1607,  669,    fmt='"$"#,##0')
ex_row(22, "Total SE tax",                           6019,     8478,  3532,   fmt='"$"#,##0', is_total=True)
ex_row(23, "AGI (W-2 + net SE − ½ SE tax)",          39591,    103761, 23234, fmt='"$"#,##0')
ex_row(24, "Standard deduction",                     15750,    31500, 15750,  fmt='"$"#,##0')
ex_row(25, "Taxable income",                         23841,    72261, 7484,   fmt='"$"#,##0')
ex_row(26, "Federal income tax (bracket calc)",      2622,     8195,  748,    fmt='"$"#,##0', is_total=True)
ex_row(27, "State income tax",                       1430,     3613,  0,      fmt='"$"#,##0', is_total=True)

# Section C: Quarterly payments
ws.row_dimensions[29].height = 24
cell(ws, "B29", "What you'd pay each quarter", font=F_H1, align=A_LEFT)
ws.merge_cells("B29:E29")

ex_row(30, "Total federal owed (SE + income tax)",   8641,     16673, 4280,   fmt='"$"#,##0')
ex_row(31, "Less: W-2 federal withholding",          0,        -5400, 0,      fmt='"$"#,##0')
ex_row(32, "Net federal owed (quarterly target ÷ 4)", 8641,    11273, 4280,   fmt='"$"#,##0', is_total=True)
ex_row(33, "Each federal quarterly payment",          2160,    2818,  1070,   fmt='"$"#,##0', is_total=True)
ex_row(34, "Each state quarterly payment",            358,     903,   0,      fmt='"$"#,##0', is_total=True)

# Section D: Safe harbor target
ws.row_dimensions[36].height = 24
cell(ws, "B36", "Safe Harbor target (whichever is lower wins)", font=F_H1, align=A_LEFT)
ws.merge_cells("B36:E36")

ex_row(37, "100% of last year's tax (110% if AGI > $150K)", 3200, 8900, 1800,  fmt='"$"#,##0')
ex_row(38, "90% of this year's projected tax",       7777,     15006, 3852,   fmt='"$"#,##0')
ex_row(39, "Easier safe harbor — total to pay across 4 quarters", 3200, 8900, 1800, fmt='"$"#,##0', is_total=True)
ex_row(40, "= safe-harbor quarterly amount",          800,     2225,  450,    fmt='"$"#,##0', is_total=True)

# Footer note
ws.row_dimensions[42].height = 70
note = ("Notice the difference: paying TRUE this-year tax is the conservative play (more cash out now, no April surprise). "
        "Paying SAFE-HARBOR is the minimum to avoid the penalty (lower quarterly checks, but you'll owe a balance in April). "
        "Both are valid. Most people pick safe-harbor for cash-flow reasons. The Quarterly Payments tab targets the TRUE number; "
        "drop it down to the safe-harbor number if cash is tight.")
cell(ws, "B42", note, font=F_BODY, align=A_LEFT_T)
ws.merge_cells("B42:E42")

# ════════════════════════════════════════════════════════════════════
# REORDER TABS — put Dashboard right after Start Here, Examples before IRS Reference
# ════════════════════════════════════════════════════════════════════
desired_order = [
    "Start Here", "Dashboard", "Setup", "Income Forecast", "Tax Calculator",
    "Quarterly Payments", "Safe Harbor", "Examples", "IRS Reference"
]
wb._sheets = [wb[s] for s in desired_order]

# ════════════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════════════
out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "quarterly-tax-system-2026.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Size: {os.path.getsize(out_path)/1024:.1f} KB")
print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
